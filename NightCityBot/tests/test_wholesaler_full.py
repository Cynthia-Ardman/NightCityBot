"""Wholesaler cog tests — helper methods, sheet parsing, state management."""

import asyncio
import json
import random
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import discord
import pytest
from openpyxl import Workbook

from NightCityBot.cogs.guns_shop import GunsShopCog as WholesalerCog


def _make_cog(tmp_path: Path, monkeypatch):
    monkeypatch.setattr("config.UNBELIEVABOAT_API_TOKEN", "test-token")
    monkeypatch.setattr("config.GUILD_ID", 111)
    monkeypatch.setattr("config.WHOLESALER_ADMIN_ROLE_IDS", "900")
    monkeypatch.setattr("config.WHOLESALER_STORE_ROLE_IDS", "800")
    monkeypatch.setattr("config.WHOLESALER_AUDIT_CHANNEL_ID", 0)
    monkeypatch.setattr("config.WHOLESALER_GOOGLE_SHEET_XLSX_URL", "")
    monkeypatch.setattr("config.WHOLESALER_XLSX_PATH", "")
    monkeypatch.setattr("config.WHOLESALER_MASTER_SHEET_NAME", "Master Gun List")

    cog = WholesalerCog.__new__(WholesalerCog)
    cog.bot = MagicMock()
    cog.bot.get_cog = MagicMock(return_value=None)
    cog.unbelievaboat = MagicMock()
    cog.lock = asyncio.Lock()
    cog.data_dir = tmp_path
    cog.state_file = tmp_path / "state.json"
    cog.store_state_file = tmp_path / "stores.json"
    cog.wholesale_inventory_file = tmp_path / "inventory" / "wholesale.json"
    cog.store_inventory_dir = tmp_path / "inventory" / "stores"
    cog.tx_file = tmp_path / "transactions.json"
    cog.sheet_cache_path = tmp_path / "master_sheet_latest.xlsx"
    cog.DEFAULT_RESTOCK_SETTINGS = WholesalerCog.DEFAULT_RESTOCK_SETTINGS
    cog.WEAPON_TYPES = WholesalerCog.WEAPON_TYPES
    cog.WEAPON_TYPE_PATTERNS = WholesalerCog.WEAPON_TYPE_PATTERNS
    cog.VALID_RESTRICTIONS = WholesalerCog.VALID_RESTRICTIONS

    cog._audit_send = AsyncMock()

    monkeypatch.setattr("NightCityBot.cogs.guns_shop.wh_lots_get_all", AsyncMock(return_value=[]))
    monkeypatch.setattr("NightCityBot.cogs.guns_shop.wh_stores_get_all", AsyncMock(return_value={}))
    monkeypatch.setattr("NightCityBot.cogs.guns_shop.wh_shops_get_all", AsyncMock(return_value={}))
    monkeypatch.setattr("NightCityBot.cogs.guns_shop.wh_settings_get", AsyncMock(return_value={}))
    monkeypatch.setattr("NightCityBot.cogs.guns_shop.wh_lots_replace_all", AsyncMock(return_value=True))
    monkeypatch.setattr("NightCityBot.cogs.guns_shop.wh_stores_replace_all", AsyncMock(return_value=True))
    monkeypatch.setattr("NightCityBot.cogs.guns_shop.wh_shops_replace_all", AsyncMock(return_value=True))
    monkeypatch.setattr("NightCityBot.cogs.guns_shop.wh_settings_save", AsyncMock(return_value=True))
    monkeypatch.setattr("NightCityBot.cogs.guns_shop.wh_tx_append", AsyncMock(return_value=True))
    monkeypatch.setattr("NightCityBot.cogs.guns_shop.pi_add_item", AsyncMock(return_value=True))

    return cog


def _make_xlsx(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Gun List"
    ws.append(["Shop A", "Gun Name", "Type", "Mag Size", "Price", "Cyberware Needed"])
    ws.append([None, "Tamayura", "Power (M)", 8, 1200, None])
    ws.append([None, "Slaught-O-Matic", "Power (L)", 36, 100, None])
    ws.append([None, "Liberty", "Power (M)", 14, 2000, None])
    ws.append([None, "M-10AF Lexington", "Power (L)", 20, 1000, None])
    ws.append([None, "Nue", "Power (M)", 10, 1300, None])
    ws.append([None, "Revolvers", "Type", None, "Price New", None])
    ws.append([None, "DR-5 Nova", "Power (H)", 6, 4000, None])
    ws.append([None, "Overture", "Power (M)", 6, 3000, None])
    ws.append([None, "Submachine Guns", "Type", None, "Price New", None])
    ws.append([None, "Guillotine", "Power (L)", 30, 3000, None])
    ws.append([None, "Shigure", "Power (L)", 24, 5000, None])
    ws.append([None, "Shotguns", "Type", None, "Price New", None])
    ws.append([None, "Crusher (6 pellets) (Auto)", "Power (L)", 8, 6000, None])
    ws.append([None, "Assault Rifles", "Type", None, "Price New", None])
    ws.append([None, "D5 Copperhead", "Power (L)", 30, 4000, None])
    ws.append([None, "Kyubi", "Power (M)", 25, 6000, None])
    xlsx = tmp_path / "guns.xlsx"
    wb.save(xlsx)
    return xlsx


def _member(member_id, is_admin=False, role_id=0):
    member = MagicMock(spec=discord.Member)
    member.id = member_id
    member.mention = f"<@{member_id}>"
    member.display_name = f"User{member_id}"
    perms = MagicMock()
    perms.administrator = is_admin
    member.guild_permissions = perms
    role = MagicMock()
    role.id = role_id
    member.roles = [role]
    return member


def _admin():
    return _member(1001, is_admin=True, role_id=900)


def _store_owner():
    return _member(2001, is_admin=False, role_id=800)


def _buyer():
    return _member(3001, is_admin=False, role_id=0)


class TestSheetParsing:
    def test_section_headers_assign_weapon_types(self, tmp_path):
        xlsx = _make_xlsx(tmp_path)
        guns = WholesalerCog.parse_master_sheet(xlsx, "Master Gun List")
        types_by_name = {g["gun_name"]: g["weapon_type"] for g in guns}
        assert types_by_name["Tamayura"] == "pistol"
        assert types_by_name["Nue"] == "pistol"
        assert types_by_name["DR-5 Nova"] == "revolver"
        assert types_by_name["Overture"] == "revolver"
        assert types_by_name["Guillotine"] == "submachine_gun"
        assert types_by_name["Crusher (6 pellets) (Auto)"] == "shotgun"
        assert types_by_name["D5 Copperhead"] == "assault_rifle"
        assert types_by_name["Kyubi"] == "assault_rifle"

    def test_section_headers_are_not_included_as_guns(self, tmp_path):
        xlsx = _make_xlsx(tmp_path)
        guns = WholesalerCog.parse_master_sheet(xlsx, "Master Gun List")
        names = {g["gun_name"] for g in guns}
        assert "Revolvers" not in names
        assert "Submachine Guns" not in names
        assert "Shotguns" not in names
        assert "Assault Rifles" not in names

    def test_all_guns_have_valid_levels(self, tmp_path):
        xlsx = _make_xlsx(tmp_path)
        guns = WholesalerCog.parse_master_sheet(xlsx, "Master Gun List")
        for g in guns:
            assert g["gun_level"] in ("L", "M", "H"), f"{g['gun_name']} has invalid level {g['gun_level']}"

    def test_all_guns_have_positive_prices(self, tmp_path):
        xlsx = _make_xlsx(tmp_path)
        guns = WholesalerCog.parse_master_sheet(xlsx, "Master Gun List")
        for g in guns:
            assert g["price_new"] > 0, f"{g['gun_name']} has price {g['price_new']}"

    def test_restriction_column_parsed_from_spreadsheet(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Master Gun List"
        ws.append(["Gun Name", "Type", "Mag Size", "Price", "Cyberware", "Restriction"])
        ws.append(["Basic Gun", "Power (L)", 10, 500, "", "basic"])
        ws.append(["Controlled Gun", "Power (M)", 10, 1000, "", "controlled"])
        ws.append(["Restricted Gun", "Power (H)", 10, 5000, "", "restricted"])
        ws.append(["No Restriction", "Power (L)", 10, 300, "", ""])
        ws.append(["Invalid Restriction", "Power (L)", 10, 400, "", "superspecial"])
        xlsx = tmp_path / "restriction_test.xlsx"
        wb.save(xlsx)
        guns = WholesalerCog.parse_master_sheet(xlsx, "Master Gun List")
        assert len(guns) == 5
        by_name = {g["gun_name"]: g for g in guns}
        assert by_name["Basic Gun"]["restriction"] == "basic"
        assert by_name["Controlled Gun"]["restriction"] == "controlled"
        assert by_name["Restricted Gun"]["restriction"] == "restricted"
        assert by_name["No Restriction"]["restriction"] == "basic"
        assert by_name["Invalid Restriction"]["restriction"] == "basic"

    def test_no_restriction_column_defaults_to_basic(self, tmp_path):
        xlsx = _make_xlsx(tmp_path)
        guns = WholesalerCog.parse_master_sheet(xlsx, "Master Gun List")
        for g in guns:
            assert g["restriction"] == "basic", f"{g['gun_name']} should default to basic"

    def test_zero_price_guns_are_filtered(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Master Gun List"
        ws.append(["Gun Name", "Type", "Mag Size", "Price", "Cyberware"])
        ws.append(["Good Gun", "Power (L)", 10, 500, ""])
        ws.append(["Free Gun", "Power (M)", 10, 0, ""])
        ws.append(["No Price", "Power (H)", 10, None, ""])
        xlsx = tmp_path / "prices.xlsx"
        wb.save(xlsx)
        guns = WholesalerCog.parse_master_sheet(xlsx, "Master Gun List")
        assert len(guns) == 1
        assert guns[0]["gun_name"] == "Good Gun"


class TestRestockLotConsolidation:
    def test_duplicate_guns_are_merged_into_single_lot(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)
        guns = [{"gun_name": "Nue", "gun_level": "M", "price_new": 1300, "weapon_type": "pistol"}]
        cfg = {
            "total_lots": 50,
            "lots_L": 0, "lots_M": 5, "lots_H": 0,
            "qty_min_L": 1, "qty_max_L": 1,
            "qty_min_M": 2, "qty_max_M": 2,
            "qty_min_H": 1, "qty_max_H": 1,
        }
        lots, totals = cog._generate_restock_lots(guns, cfg, random.Random(42))
        assert len(lots) == 1
        assert lots[0]["gun_name"] == "Nue"
        assert lots[0]["qty_available"] == 10
        assert totals["M"] == 10

    def test_different_guns_stay_separate(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)
        guns = [
            {"gun_name": "Nue", "gun_level": "M", "price_new": 1300, "weapon_type": "pistol"},
            {"gun_name": "Liberty", "gun_level": "M", "price_new": 2000, "weapon_type": "pistol"},
        ]
        cfg = {
            "total_lots": 50,
            "lots_L": 0, "lots_M": 4, "lots_H": 0,
            "qty_min_L": 1, "qty_max_L": 1,
            "qty_min_M": 1, "qty_max_M": 1,
            "qty_min_H": 1, "qty_max_H": 1,
        }
        lots, _ = cog._generate_restock_lots(guns, cfg, random.Random(42))
        assert len(lots) <= 2
        for lot in lots:
            assert lot["weapon_type"] == "pistol"

    def test_same_gun_different_levels_stay_separate(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)
        guns = [
            {"gun_name": "TestGun", "gun_level": "L", "price_new": 100, "weapon_type": "pistol"},
            {"gun_name": "TestGun", "gun_level": "H", "price_new": 500, "weapon_type": "pistol"},
        ]
        cfg = {
            "total_lots": 50,
            "lots_L": 2, "lots_M": 0, "lots_H": 2,
            "qty_min_L": 1, "qty_max_L": 1,
            "qty_min_M": 1, "qty_max_M": 1,
            "qty_min_H": 1, "qty_max_H": 1,
        }
        lots, _ = cog._generate_restock_lots(guns, cfg, random.Random(42))
        assert len(lots) == 2
        levels = {l["gun_level"] for l in lots}
        assert levels == {"L", "H"}


class TestStateManagement:
    def test_save_and_load_roundtrip(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)
        state = {
            "wholesale_lots": [
                {"lot_id": "lot-001", "gun_name": "Nue", "gun_level": "M",
                 "weapon_type": "pistol", "unit_cost": 1300, "qty_available": 5}
            ],
            "stores": {
                "111:2001": {
                    "owner_id": 2001,
                    "lots": [
                        {"lot_id": "lot-001", "gun_name": "Nue", "gun_level": "M",
                         "weapon_type": "pistol", "unit_cost": 1300, "qty_remaining": 3}
                    ]
                }
            },
            "shop_registry": {"shop1": 2001},
            "settings": {},
        }

        async def _run():
            ok = await cog._save_state(state)
            assert ok is True
            loaded = await cog._load_state()
            return loaded

        loaded = asyncio.run(_run())
        assert loaded["wholesale_lots"][0]["lot_id"] == "lot-001"
        assert loaded["wholesale_lots"][0]["weapon_type"] == "pistol"
        assert loaded["stores"]["111:2001"]["lots"][0]["qty_remaining"] == 3
        assert loaded["shop_registry"]["shop1"] == 2001

    def test_empty_state_initializes_defaults(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)

        async def _run():
            return await cog._load_state()

        state = asyncio.run(_run())
        assert "wholesale_lots" in state
        assert isinstance(state["wholesale_lots"], list)

    def test_save_creates_per_store_inventory_files(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)
        state = {
            "wholesale_lots": [],
            "stores": {
                "111:2001": {
                    "owner_id": 2001,
                    "lots": [{"lot_id": "s1", "gun_name": "Nue", "qty_remaining": 3}]
                },
                "111:2002": {
                    "owner_id": 2002,
                    "lots": [{"lot_id": "s2", "gun_name": "Liberty", "qty_remaining": 1}]
                },
            },
            "shop_registry": {},
            "settings": {},
        }

        async def _run():
            await cog._save_state(state)

        asyncio.run(_run())
        assert cog.store_inventory_dir.exists()
        store_files = list(cog.store_inventory_dir.glob("*.json"))
        assert len(store_files) == 2


class TestPermissionHelpers:
    def test_admin_check_accepts_administrator_permission(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)
        admin = _admin()
        assert cog._is_admin(admin) is True

    def test_admin_check_rejects_regular_user(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)
        person = _buyer()
        assert cog._is_admin(person) is False

    def test_store_owner_check_with_role(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)
        owner = _store_owner()
        assert cog._is_store_owner(owner) is True

    def test_store_owner_check_without_role(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)
        person = _buyer()
        assert cog._is_store_owner(person) is False

    def test_admin_does_not_bypass_store_owner_check(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)
        admin = _member(1001, is_admin=True, role_id=900)
        assert cog._is_store_owner(admin) is (900 in cog._coerce_role_ids("800"))


class TestEdgeCases:
    def test_weapon_type_derivation_patterns(self, tmp_path, monkeypatch):
        assert WholesalerCog._derive_weapon_type("Crusher", "Power Revolver (L)") == "revolver"
        assert WholesalerCog._derive_weapon_type("M221 Saratoga", "SMG (M)") == "submachine_gun"
        assert WholesalerCog._derive_weapon_type("TestGun", "Power (L)") is None
        assert WholesalerCog._derive_weapon_type("Test Pistol", "Power (L)") == "pistol"
        assert WholesalerCog._derive_weapon_type("Test Shotgun", "Power (L)") == "shotgun"
        assert WholesalerCog._derive_weapon_type("Test Gun", "Assault Rifle (M)") == "assault_rifle"
        assert WholesalerCog._derive_weapon_type("LMG Test", "Light Machine Gun (H)") == "light_machine_gun"
        assert WholesalerCog._derive_weapon_type("Sniper X", "Sniper Rifle (H)") == "sniper_rifle"

    def test_level_derivation_edge_cases(self):
        assert WholesalerCog._derive_level("Tech (M-H)") == "H"
        assert WholesalerCog._derive_level("Tech (M)") == "M"
        assert WholesalerCog._derive_level("Power (L)") == "L"
        assert WholesalerCog._derive_level("Smart (H)") == "H"

    def test_normalize_shop_name(self):
        assert WholesalerCog._normalize_shop_name("Shop 1") == "shop-1"
        assert WholesalerCog._normalize_shop_name("SHOP__2") == "shop-2"

    def test_store_id_generation(self):
        assert WholesalerCog._store_id(111, 2001) == "111:2001"
        assert WholesalerCog._store_id(999, 1) == "999:1"

    def test_google_sheet_url_normalization(self):
        url = "https://docs.google.com/spreadsheets/d/abc123/edit?usp=sharing"
        out = WholesalerCog._normalize_sheet_source_url(url)
        assert out == "https://docs.google.com/spreadsheets/d/abc123/export?format=xlsx"

    def test_google_sheet_url_with_gid(self):
        url = "https://docs.google.com/spreadsheets/d/abc123/edit#gid=456"
        out = WholesalerCog._normalize_sheet_source_url(url)
        assert out == "https://docs.google.com/spreadsheets/d/abc123/export?format=xlsx&gid=456"

    def test_google_sheet_url_strips_angle_brackets(self):
        url = "<https://docs.google.com/spreadsheets/d/abc123/edit?usp=sharing>"
        out = WholesalerCog._normalize_sheet_source_url(url)
        assert out == "https://docs.google.com/spreadsheets/d/abc123/export?format=xlsx"

    def test_restock_with_zero_level_lots(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)
        cfg = cog._resolve_restock_settings({
            "settings": {
                "restock": {
                    "total_lots": 5,
                    "lots_L": 5, "lots_M": 0, "lots_H": 0,
                    "qty_min_L": 2, "qty_max_L": 2,
                }
            }
        })
        assert cfg["lots_M"] == 0
        assert cfg["lots_H"] == 0
        assert cfg["total_lots"] == 5

    def test_restock_defaults_restriction_to_basic(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)
        guns = [
            {"gun_name": "Nue", "gun_level": "M", "price_new": 1300, "weapon_type": "pistol"},
        ]
        cfg = {
            "total_lots": 50,
            "lots_L": 0, "lots_M": 1, "lots_H": 0,
            "qty_min_L": 1, "qty_max_L": 1,
            "qty_min_M": 1, "qty_max_M": 1,
            "qty_min_H": 1, "qty_max_H": 1,
        }
        lots, _totals = cog._generate_restock_lots(guns, cfg, random.Random(42))
        assert len(lots) == 1
        assert lots[0]["restriction"] == "basic"

    def test_restock_carries_restriction_from_parsed_guns(self, tmp_path, monkeypatch):
        cog = _make_cog(tmp_path, monkeypatch)
        guns = [
            {"gun_name": "Nue", "gun_level": "M", "price_new": 1300, "weapon_type": "pistol", "restriction": "controlled"},
        ]
        cfg = {
            "total_lots": 50,
            "lots_L": 0, "lots_M": 3, "lots_H": 0,
            "qty_min_L": 1, "qty_max_L": 1,
            "qty_min_M": 1, "qty_max_M": 1,
            "qty_min_H": 1, "qty_max_H": 1,
        }
        lots, _totals = cog._generate_restock_lots(guns, cfg, random.Random(42))
        assert len(lots) == 1
        assert lots[0]["restriction"] == "controlled"
