import asyncio
import logging
import random
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional
from urllib.parse import urlparse, parse_qs

import aiohttp
import discord
from discord.ext import commands, tasks
from openpyxl import load_workbook

import config
from NightCityBot.services.unbelievaboat import UnbelievaBoatAPI
from NightCityBot.utils import helpers

logger = logging.getLogger(__name__)


class WholesalerCog(commands.Cog):
    """Two-tier gun supply chain: corp wholesaler -> stores -> players.

    Uses read-only spreadsheet parsing and immutable receipt/audit logs.
    Staff updates Character Gun Tracking manually from receipts.
    """

    LEVEL_SETTINGS = {
        "L": {"weight": 70, "qty_min": 3, "qty_max": 10},
        "M": {"weight": 25, "qty_min": 1, "qty_max": 5},
        "H": {"weight": 5, "qty_min": 1, "qty_max": 2},
    }
    DEFAULT_RESTOCK_SETTINGS = {
        "total_lots": 20,
        "lots_L": 14,
        "lots_M": 5,
        "lots_H": 1,
        "qty_min_L": 3,
        "qty_max_L": 10,
        "qty_min_M": 1,
        "qty_max_M": 5,
        "qty_min_H": 1,
        "qty_max_H": 2,
    }
    WEAPON_TYPES = (
        "revolver",
        "submachine_gun",
        "shotgun",
        "assault_rifle",
        "light_machine_gun",
        "heavy_machine_gun",
        "precision_rifle",
        "sniper_rifle",
    )
    WEAPON_TYPE_PATTERNS = {
        "revolver": ("revolver",),
        "submachine_gun": ("submachine gun", "submachine-gun", "smg"),
        "shotgun": ("shotgun",),
        "assault_rifle": ("assault rifle", "assault-rifle", "ar "),
        "light_machine_gun": ("light machine gun", "light-machine-gun", "lmg"),
        "heavy_machine_gun": ("heavy machine gun", "heavy-machine-gun", "hmg"),
        "precision_rifle": ("precision rifle", "precision-rifle", "dmr"),
        "sniper_rifle": ("sniper rifle", "sniper-rifle", "sniper"),
    }

    def __init__(self, bot: commands.Bot) -> None:
        self.bot = bot
        self.unbelievaboat = UnbelievaBoatAPI(config.UNBELIEVABOAT_API_TOKEN)

        base_dir = Path(getattr(config, "BASE_DIR", Path(__file__).resolve().parents[1]))
        self.base_dir = base_dir
        default_data_dir = base_dir / "data" / "wholesaler"
        configured_data_dir = getattr(config, "WHOLESALER_DATA_DIR", None)
        self.data_dir = self._resolve_base_path(configured_data_dir, default_data_dir)
        self.state_file = self._resolve_data_path(
            getattr(config, "WHOLESALER_STATE_FILE", None),
            "state.json",
        )

        self.data_dir.mkdir(parents=True, exist_ok=True)

        self.sheet_cache_path = self._resolve_data_path(
            getattr(config, "WHOLESALER_SHEET_CACHE_FILE", None),
            "master_sheet_latest.xlsx",
        )
        self.store_state_file = self._resolve_data_path(
            getattr(config, "WHOLESALER_STORE_STATE_FILE", None),
            "stores.json",
        )
        self.wholesale_inventory_file = self._resolve_data_path(
            getattr(config, "WHOLESALER_WHOLESALE_FILE", None),
            "inventory/wholesale.json",
        )
        self.store_inventory_dir = self._resolve_data_path(
            getattr(config, "WHOLESALER_STORE_INVENTORY_DIR", None),
            "inventory/stores",
        )
        self.tx_file = self._resolve_data_path(
            getattr(config, "WHOLESALER_TX_FILE", None),
            "transactions.json",
        )

        self._migrate_legacy_files(default_data_dir)
        self.wholesale_inventory_file.parent.mkdir(parents=True, exist_ok=True)
        self.store_inventory_dir.mkdir(parents=True, exist_ok=True)
        self.lock = asyncio.Lock()
        self._startup_audit_sent = False
        self.weekly_sunday_restock.start()

    def _resolve_data_path(self, configured_path: Any, default_relative: str) -> Path:
        if configured_path is None:
            return self.data_dir / default_relative

        configured = Path(str(configured_path)).expanduser()
        if configured.is_absolute():
            return configured
        return self.data_dir / configured

    def _resolve_base_path(self, configured_path: Any, default_path: Path) -> Path:
        """Resolve base-level paths, anchoring relative values to config.BASE_DIR."""
        if configured_path is None:
            return default_path

        configured = Path(str(configured_path)).expanduser()
        if configured.is_absolute():
            return configured
        return self.base_dir / configured

    def _migrate_legacy_files(self, legacy_dir: Path) -> None:
        """Copy old in-repo wholesaler files into configured persistent paths once."""
        if self.data_dir == legacy_dir:
            return

        store_state_file = getattr(self, "store_state_file", self.state_file)

        migrations = (
            (legacy_dir / "state.json", self.state_file),
            (legacy_dir / "state.json", store_state_file),
            (legacy_dir / "transactions.json", self.tx_file),
            (legacy_dir / "master_sheet_latest.xlsx", self.sheet_cache_path),
        )
        for src, dst in migrations:
            if dst.exists() or not src.exists():
                continue
            dst.parent.mkdir(parents=True, exist_ok=True)
            try:
                dst.write_bytes(src.read_bytes())
                logger.info("Migrated wholesaler data file %s -> %s", src, dst)
            except Exception:
                logger.exception("Failed to migrate wholesaler data file %s -> %s", src, dst)

        legacy_wholesale = legacy_dir / "inventory" / "wholesale.json"
        if legacy_wholesale.exists() and not self.wholesale_inventory_file.exists():
            self.wholesale_inventory_file.parent.mkdir(parents=True, exist_ok=True)
            try:
                self.wholesale_inventory_file.write_bytes(legacy_wholesale.read_bytes())
                logger.info("Migrated wholesaler inventory file %s -> %s", legacy_wholesale, self.wholesale_inventory_file)
            except Exception:
                logger.exception("Failed to migrate wholesaler inventory file %s -> %s", legacy_wholesale, self.wholesale_inventory_file)

        legacy_store_dir = legacy_dir / "inventory" / "stores"
        if legacy_store_dir.exists():
            self.store_inventory_dir.mkdir(parents=True, exist_ok=True)
            for src in sorted(legacy_store_dir.glob("*.json")):
                dst = self.store_inventory_dir / src.name
                if dst.exists():
                    continue
                try:
                    dst.write_bytes(src.read_bytes())
                    logger.info("Migrated store inventory file %s -> %s", src, dst)
                except Exception:
                    logger.exception("Failed to migrate store inventory file %s -> %s", src, dst)


    def _store_inventory_file(self, store_id: str) -> Path:
        safe_store_id = re.sub(r"[^a-zA-Z0-9_-]+", "-", str(store_id)).strip("-")
        store_inventory_dir = Path(getattr(self, "store_inventory_dir", Path(getattr(self, "state_file", Path("state.json"))).parent / "inventory" / "stores"))
        return store_inventory_dir / f"{safe_store_id or 'store'}.json"

    def _list_store_inventory_files(self) -> list[Path]:
        store_inventory_dir = Path(getattr(self, "store_inventory_dir", Path(getattr(self, "state_file", Path("state.json"))).parent / "inventory" / "stores"))
        if not store_inventory_dir.exists():
            return []
        return sorted(store_inventory_dir.glob("*.json"))

    def cog_unload(self):
        self.weekly_sunday_restock.cancel()
        self.bot.loop.create_task(self.unbelievaboat.close())

    @tasks.loop(hours=1)
    async def weekly_sunday_restock(self) -> None:
        """Automatically refresh wholesaler stock every Sunday (UTC)."""
        now = datetime.now(timezone.utc)
        if now.weekday() != 6:
            return
        await self._auto_restock_if_due(now, trigger="SCHEDULED")

    @weekly_sunday_restock.before_loop
    async def before_weekly_sunday_restock(self) -> None:
        await self.bot.wait_until_ready()

    @staticmethod
    def _now_iso() -> str:
        return datetime.now(timezone.utc).isoformat()

    @staticmethod
    def _to_int(value: Any) -> Optional[int]:
        if value is None:
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        stripped = str(value).replace(",", "").replace("$", "").strip()
        if not stripped:
            return None
        try:
            return int(float(stripped))
        except ValueError:
            return None

    @staticmethod
    def _derive_level(effectiveness_raw: str) -> str:
        text = (effectiveness_raw or "").upper()
        if "(M-H)" in text or "(H)" in text:
            return "H"
        if "(M)" in text:
            return "M"
        if "(L)" in text:
            return "L"
        return "L"

    @staticmethod
    def _derive_category(effectiveness_raw: str) -> Optional[str]:
        text = (effectiveness_raw or "").lower()
        for category in ("power", "tech", "smart"):
            if category in text:
                return category.title()
        return None

    @classmethod
    def _derive_weapon_type(cls, gun_name: str, effectiveness_raw: str) -> Optional[str]:
        haystack = f"{gun_name or ''} {effectiveness_raw or ''}".lower()
        for weapon_type, patterns in cls.WEAPON_TYPE_PATTERNS.items():
            if any(pattern in haystack for pattern in patterns):
                return weapon_type
        return None

    @staticmethod
    def _normalize_shop_name(name: str) -> str:
        cleaned = re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-")
        return cleaned

    @staticmethod
    def _inventory_totals(state: dict[str, Any]) -> tuple[int, int, int, int]:
        wholesale_lots = state.get("wholesale_lots", [])
        wholesale_lot_count = len(wholesale_lots) if isinstance(wholesale_lots, list) else 0
        wholesale_unit_count = (
            sum(max(int(lot.get("qty_available", 0)), 0) for lot in wholesale_lots)
            if isinstance(wholesale_lots, list)
            else 0
        )

        stores = state.get("stores", {})
        store_count = len(stores) if isinstance(stores, dict) else 0
        store_unit_count = 0
        if isinstance(stores, dict):
            for payload in stores.values():
                if not isinstance(payload, dict):
                    continue
                for lot in payload.get("lots", []):
                    if isinstance(lot, dict):
                        store_unit_count += max(int(lot.get("qty_remaining", 0)), 0)

        return wholesale_lot_count, wholesale_unit_count, store_count, store_unit_count

    async def emit_inventory_snapshot_audit(
        self,
        reason: str,
        *,
        actor: Optional[discord.abc.User] = None,
    ) -> None:
        """Emit an audit line that captures persisted inventory state."""
        try:
            state = await self._load_state()
            lots, wholesale_units, stores, store_units = self._inventory_totals(state)
            wholesale_exists = self.wholesale_inventory_file.exists()
            wholesale_bytes = self.wholesale_inventory_file.stat().st_size if wholesale_exists else 0
            state_exists = self.state_file.exists()
            state_bytes = self.state_file.stat().st_size if state_exists else 0
            actor_text = f" actor=<@{actor.id}>" if actor else ""
            await self._audit_send(
                "[WHOLESALE_SNAPSHOT]"
                f" reason={reason}{actor_text}"
                f" lots={lots} wholesale_units={wholesale_units}"
                f" stores={stores} store_units={store_units}"
                f" wholesale_file={self.wholesale_inventory_file}"
                f" wholesale_file_exists={wholesale_exists} wholesale_file_bytes={wholesale_bytes}"
                f" state_file={self.state_file}"
                f" state_file_exists={state_exists} state_file_bytes={state_bytes}"
            )
        except Exception:
            logger.exception("Failed to emit wholesaler inventory snapshot for reason=%s", reason)

    @commands.Cog.listener()
    async def on_ready(self) -> None:
        if self._startup_audit_sent:
            return
        self._startup_audit_sent = True
        logger.info(
            "Wholesaler data paths: data_dir=%s state=%s wholesale=%s stores=%s store_inv_dir=%s tx=%s",
            self.data_dir,
            self.state_file,
            self.wholesale_inventory_file,
            self.store_state_file,
            self.store_inventory_dir,
            self.tx_file,
        )
        await self._ensure_inventory_files_exist()
        logger.info(
            "Wholesaler files verified: state=%s wholesale=%s stores=%s tx=%s",
            self.state_file.exists(),
            self.wholesale_inventory_file.exists(),
            self.store_state_file.exists(),
            self.tx_file.exists(),
        )
        await self.emit_inventory_snapshot_audit("BOT_READY")

    async def _ensure_inventory_files_exist(self) -> None:
        """Create baseline wholesaler/store persistence files during startup."""
        async with self.lock:
            data_dir = Path(getattr(self, "data_dir", self.state_file.parent))
            # Create expected files/directories up front so deployments can
            # validate storage locations immediately after startup.
            data_dir.mkdir(parents=True, exist_ok=True)
            self.wholesale_inventory_file.parent.mkdir(parents=True, exist_ok=True)
            self.store_inventory_dir.mkdir(parents=True, exist_ok=True)

            if not self.state_file.exists():
                await helpers.save_json_file(
                    self.state_file,
                    {
                        "wholesale_lots": [],
                        "transactions": 0,
                        "pending_payouts": [],
                        "settings": {},
                    },
                )
            if not self.store_state_file.exists():
                await helpers.save_json_file(
                    self.store_state_file,
                    {
                        "shop_registry": {},
                        "stores": {},
                    },
                )
            if not self.wholesale_inventory_file.exists():
                await helpers.save_json_file(
                    self.wholesale_inventory_file,
                    {
                        "wholesale_lots": [],
                    },
                )
            if not self.tx_file.exists():
                await helpers.save_json_file(self.tx_file, [])

            state = await self._load_state()
            saved = await self._save_state(state)
            if not saved:
                logger.warning(
                    "Wholesaler startup bootstrap could not persist one or more inventory files "
                    "(state=%s, wholesale=%s, store_index=%s, store_dir=%s)",
                    self.state_file,
                    self.wholesale_inventory_file,
                    self.store_state_file,
                    self.store_inventory_dir,
                )
            base_dir = Path(getattr(self, "base_dir", data_dir.parent))
            if data_dir != base_dir / "data" / "wholesaler":
                logger.info(
                    "Wholesaler data directory override active data_dir=%s base_dir=%s",
                    data_dir,
                    base_dir,
                )


    @staticmethod
    def _clean_sheet_url_input(value: str) -> str:
        """Normalize user-provided URL text from Discord messages."""
        raw = (value or "").strip()
        if raw.startswith("<") and raw.endswith(">"):
            raw = raw[1:-1].strip()
        return raw

    @staticmethod
    def _normalize_sheet_source_url(value: str) -> str:
        """Normalize a Google Sheets URL to an XLSX export URL when possible."""
        raw = WholesalerCog._clean_sheet_url_input(value)
        if not raw:
            return raw
        parsed = urlparse(raw)

        if "docs.google.com" not in parsed.netloc:
            return raw

        # Already an export endpoint
        if "/export" in parsed.path:
            return raw

        match = re.search(r"/spreadsheets/(?:u/\d+/)?d/([a-zA-Z0-9-_]+)", parsed.path)
        if not match:
            return raw

        sheet_id = match.group(1)
        gid = parse_qs(parsed.query).get("gid", [None])[0]
        if not gid and parsed.fragment:
            frag_qs = parse_qs(parsed.fragment)
            gid = frag_qs.get("gid", [None])[0]
            if not gid and parsed.fragment.startswith("gid="):
                gid = parsed.fragment.split("=", 1)[1].split("&", 1)[0]

        export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        if gid:
            export_url += f"&gid={gid}"
        return export_url

    @staticmethod
    def _extract_sheet_gid(value: str) -> Optional[str]:
        """Extract a Google Sheets gid from a URL query or fragment."""
        raw = WholesalerCog._clean_sheet_url_input(value)
        if not raw:
            return None
        parsed = urlparse(raw)
        gid = parse_qs(parsed.query).get("gid", [None])[0]
        if gid:
            return gid
        if parsed.fragment:
            frag_qs = parse_qs(parsed.fragment)
            gid = frag_qs.get("gid", [None])[0]
            if gid:
                return gid
            if parsed.fragment.startswith("gid="):
                return parsed.fragment.split("=", 1)[1].split("&", 1)[0]
        return None

    @staticmethod
    def parse_master_sheet(
        xlsx_path: str | Path,
        sheet_name: str,
        sheet_gid: Optional[str] = None,
    ) -> list[dict[str, Any]]:
        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
        selected_sheet = sheet_name if sheet_name in wb.sheetnames else None
        if not selected_sheet and sheet_gid:
            for candidate in wb.worksheets:
                tab_id = getattr(candidate.sheet_properties, "tabId", None)
                if tab_id is not None and str(tab_id) == str(sheet_gid):
                    selected_sheet = candidate.title
                    break

        # Some exports only include one visible tab. Use it as a safe fallback.
        if not selected_sheet and len(wb.sheetnames) == 1:
            selected_sheet = wb.sheetnames[0]

        if not selected_sheet:
            available = ", ".join(wb.sheetnames)
            wb.close()
            raise ValueError(
                f"Sheet '{sheet_name}' not found in workbook"
                f" (available: {available})"
            )

        ws = wb[selected_sheet]
        row_iter = ws.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if not header_row:
            wb.close()
            return []

        header = [str(c).strip() if c is not None else "" for c in header_row]
        header_lookup = {name.lower(): idx for idx, name in enumerate(header)}

        def idx_for(options: list[str], fallback: int) -> int:
            for opt in options:
                found = header_lookup.get(opt.lower())
                if found is not None:
                    return found
            return fallback

        name_idx = idx_for(["Gun Name", "Name", "Weapon"], 0)
        eff_idx = idx_for(["Type/Armor Effectiveness", "Type", "Effectiveness"], 1)
        mag_idx = idx_for(["Mag Size", "Mag"], 2)
        price_idx = idx_for(["Price New", "Price", "Price (New)"], 3)
        cyberware_idx = idx_for(["Cyberware Needed", "Cyberware"], 4)

        parsed: list[dict[str, Any]] = []
        for row in row_iter:
            if not row:
                continue

            gun_name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] is not None else ""
            effectiveness_raw = (
                str(row[eff_idx]).strip() if eff_idx < len(row) and row[eff_idx] is not None else ""
            )
            price_new = WholesalerCog._to_int(row[price_idx] if price_idx < len(row) else None)

            if not gun_name:
                continue
            if effectiveness_raw.lower() == "type":
                continue
            if price_new is None or price_new <= 0:
                continue

            mag_raw = row[mag_idx] if mag_idx < len(row) else None
            mag_size = WholesalerCog._to_int(mag_raw)
            if mag_size is None and mag_raw is not None:
                mag_size = str(mag_raw)

            cyber_raw = row[cyberware_idx] if cyberware_idx < len(row) else ""
            cyberware_needed = WholesalerCog._to_int(cyber_raw)
            if cyberware_needed is None:
                cyberware_needed = "" if cyber_raw is None else str(cyber_raw)

            parsed.append(
                {
                    "gun_name": gun_name,
                    "effectiveness_raw": effectiveness_raw,
                    "mag_size": mag_size,
                    "price_new": price_new,
                    "cyberware_needed": cyberware_needed,
                    "gun_level": WholesalerCog._derive_level(effectiveness_raw),
                    "gun_category": WholesalerCog._derive_category(effectiveness_raw),
                    "weapon_type": WholesalerCog._derive_weapon_type(gun_name, effectiveness_raw),
                }
            )

        wb.close()
        return parsed

    async def _resolve_sheet_source(self) -> tuple[str, Optional[str]]:
        """Resolve active master-sheet URL and optional gid override."""
        state = await self._load_state()
        configured_url = str(state.get("settings", {}).get("master_sheet_url", "")).strip()
        config_url = str(getattr(config, "WHOLESALER_GOOGLE_SHEET_XLSX_URL", "") or "").strip()
        raw_url = configured_url or config_url

        local_fallback = str(getattr(config, "WHOLESALER_XLSX_PATH", "") or "").strip()
        if not raw_url and local_fallback.startswith(("http://", "https://")):
            raw_url = local_fallback

        return self._normalize_sheet_source_url(raw_url), self._extract_sheet_gid(raw_url)

    async def _resolve_sheet_path(self) -> Path:
        """Return local xlsx path, downloading from Google Sheets if configured."""
        state = await self._load_state()
        configured_url = str(state.get("settings", {}).get("master_sheet_url", "")).strip()
        sheet_url, _ = await self._resolve_sheet_source()
        if not sheet_url:
            local_path = Path(config.WHOLESALER_XLSX_PATH)
            if not local_path.exists():
                raise FileNotFoundError(
                    f"Local wholesaler sheet not found at '{local_path}'. "
                    "Set a Google Sheet source with !wh_setsheet <url>, or update WHOLESALER_XLSX_PATH."
                )
            return local_path

        # Self-heal older stored share URLs by persisting normalized export URL.
        if configured_url and configured_url != sheet_url:
            async with self.lock:
                latest = await self._load_state()
                latest.setdefault("settings", {})["master_sheet_url"] = sheet_url
                await self._save_state(latest)

        timeout = aiohttp.ClientTimeout(total=30)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(sheet_url) as resp:
                if resp.status != 200:
                    raise RuntimeError(f"Failed to fetch sheet export ({resp.status})")
                payload = await resp.read()
                await asyncio.to_thread(self.sheet_cache_path.write_bytes, payload)
        return self.sheet_cache_path

    async def _load_master_guns(self) -> list[dict[str, Any]]:
        """Load wholesaler source rows using configured sheet name with gid fallback."""
        sheet_path = await self._resolve_sheet_path()
        _, sheet_gid = await self._resolve_sheet_source()
        return await asyncio.to_thread(
            self.parse_master_sheet,
            sheet_path,
            config.WHOLESALER_MASTER_SHEET_NAME,
            sheet_gid,
        )

    async def _load_state(self) -> dict[str, Any]:
        wholesale_file = Path(getattr(self, "wholesale_inventory_file", self.state_file))
        store_file = Path(getattr(self, "store_state_file", self.state_file))
        default_data_dir = Path(getattr(self, "data_dir", Path(self.state_file).parent))
        store_inventory_dir = Path(getattr(self, "store_inventory_dir", default_data_dir / "inventory" / "stores"))

        state = await helpers.load_json_file(
            self.state_file,
            default={
                "wholesale_lots": [],
                "transactions": 0,
                "pending_payouts": [],
                "settings": {},
            },
        )

        wholesale_state = await helpers.load_json_file(wholesale_file, default={})

        stores: dict[str, Any] = {}
        shop_registry: dict[str, Any] = {}

        if store_file == self.state_file:
            store_state = state
        else:
            store_state = await helpers.load_json_file(
                store_file,
                default={"stores": {}, "shop_registry": {}},
            )

        legacy_stores = store_state.get("stores", {})
        if isinstance(legacy_stores, dict):
            for store_id, payload in legacy_stores.items():
                if not isinstance(payload, dict):
                    continue
                lots = payload.get("lots", [])
                stores[str(store_id)] = {
                    "owner_id": payload.get("owner_id"),
                    "lots": lots if isinstance(lots, list) else [],
                }

        legacy_registry = store_state.get("shop_registry", {})
        if isinstance(legacy_registry, dict):
            shop_registry.update(legacy_registry)

        if store_inventory_dir.exists():
            for store_file_path in sorted(store_inventory_dir.glob("*.json")):
                payload = await helpers.load_json_file(store_file_path, default={})
                if not isinstance(payload, dict):
                    continue
                store_id = str(payload.get("store_id") or store_file_path.stem)
                lots = payload.get("lots", [])
                if isinstance(lots, list):
                    owner_id = payload.get("owner_id")
                    stores[store_id] = {"owner_id": owner_id, "lots": lots}

        state["stores"] = stores
        state["shop_registry"] = shop_registry
        wholesale_lots = wholesale_state.get("wholesale_lots", state.get("wholesale_lots", []))
        state["wholesale_lots"] = wholesale_lots if isinstance(wholesale_lots, list) else []
        state.setdefault("shop_registry", {})
        state.setdefault("stores", {})
        state.setdefault("wholesale_lots", [])
        state.setdefault("pending_payouts", [])
        state.setdefault("settings", {})
        restock = state["settings"].setdefault("restock", {})
        for key, value in self.DEFAULT_RESTOCK_SETTINGS.items():
            restock.setdefault(key, value)
        return state

    @staticmethod
    def _sanitize_positive_int(value: Any, fallback: int) -> int:
        try:
            v = int(value)
            return v if v > 0 else fallback
        except Exception:
            return fallback

    @staticmethod
    def _sanitize_non_negative_int(value: Any, fallback: int) -> int:
        try:
            v = int(value)
            return v if v >= 0 else fallback
        except Exception:
            return fallback

    def _resolve_restock_settings(self, state: dict[str, Any]) -> dict[str, int]:
        raw = state.get("settings", {}).get("restock", {})
        data = {}
        for key, default in self.DEFAULT_RESTOCK_SETTINGS.items():
            if key in {"lots_L", "lots_M", "lots_H"}:
                data[key] = self._sanitize_non_negative_int(raw.get(key), default)
            else:
                data[key] = self._sanitize_positive_int(raw.get(key), default)

        # Keep ranges valid and ensure at least one lot is generated.
        for lvl in ("L", "M", "H"):
            mn_key = f"qty_min_{lvl}"
            mx_key = f"qty_max_{lvl}"
            if data[mn_key] > data[mx_key]:
                data[mn_key], data[mx_key] = data[mx_key], data[mn_key]

        if data["lots_L"] + data["lots_M"] + data["lots_H"] <= 0:
            data["lots_L"] = 1

        data["total_lots"] = max(1, data["total_lots"])

        explicit_type_lots = state.get("settings", {}).get("restock_type_lots", {})
        for weapon_type in self.WEAPON_TYPES:
            for lvl in ("L", "M", "H"):
                key = f"{weapon_type}_{lvl}"
                data[key] = self._sanitize_non_negative_int(explicit_type_lots.get(key), 0)

        return data

    def _generate_restock_lots(
        self,
        guns: list[dict[str, Any]],
        cfg: dict[str, int],
        rng: random.Random,
    ) -> tuple[list[dict[str, Any]], dict[str, int]]:
        by_level = {
            "L": [g for g in guns if g["gun_level"] == "L"],
            "M": [g for g in guns if g["gun_level"] == "M"],
            "H": [g for g in guns if g["gun_level"] == "H"],
        }
        weighted = [
            g
            for g in guns
            for _ in range(self.LEVEL_SETTINGS.get(g["gun_level"], {"weight": 1})["weight"])
        ]

        lots: list[dict[str, Any]] = []
        level_totals = {"L": 0, "M": 0, "H": 0}

        explicit_keys = [f"{weapon_type}_{lvl}" for weapon_type in self.WEAPON_TYPES for lvl in ("L", "M", "H")]
        has_explicit_mix = any(cfg.get(k, 0) > 0 for k in explicit_keys)

        if has_explicit_mix:
            for weapon_type in self.WEAPON_TYPES:
                for requested_level in ("L", "M", "H"):
                    target = int(cfg.get(f"{weapon_type}_{requested_level}", 0))
                    if target <= 0:
                        continue
                    pool = [
                        g for g in guns
                        if g.get("weapon_type") == weapon_type and g.get("gun_level") == requested_level
                    ]
                    if not pool:
                        pool = [g for g in by_level[requested_level] if g.get("weapon_type") == weapon_type]
                    if not pool:
                        continue

                    for _ in range(target):
                        gun = rng.choice(pool)
                        qty = rng.randint(cfg[f"qty_min_{requested_level}"], cfg[f"qty_max_{requested_level}"])
                        lots.append(
                            {
                                "lot_id": f"lot-{datetime.now(timezone.utc).strftime('%Y%m%d')}-{uuid.uuid4().hex[:6]}",
                                "gun_name": gun["gun_name"],
                                "gun_level": requested_level,
                                "unit_cost": int(gun["price_new"]),
                                "qty_available": qty,
                                "created_at": self._now_iso(),
                            }
                        )
                        level_totals[requested_level] += qty
        else:

            for requested_level in ("L", "M", "H"):
                target = cfg[f"lots_{requested_level}"]
                pool = by_level[requested_level] if by_level[requested_level] else weighted
                if not pool or target <= 0:
                    continue

                for _ in range(target):
                    gun = rng.choice(pool)
                    actual_level = str(gun.get("gun_level", requested_level))
                    if actual_level not in {"L", "M", "H"}:
                        actual_level = requested_level
                    qty = rng.randint(cfg[f"qty_min_{actual_level}"], cfg[f"qty_max_{actual_level}"])
                    lots.append(
                        {
                            "lot_id": f"lot-{datetime.now(timezone.utc).strftime('%Y%m%d')}-{uuid.uuid4().hex[:6]}",
                            "gun_name": gun["gun_name"],
                            "gun_level": actual_level,
                            "unit_cost": int(gun["price_new"]),
                            "qty_available": qty,
                            "created_at": self._now_iso(),
                        }
                    )
                    level_totals[actual_level] += qty

        if len(lots) > cfg["total_lots"]:
            lots = rng.sample(lots, cfg["total_lots"])
            level_totals = {
                "L": sum(int(lot.get("qty_available", 0)) for lot in lots if lot.get("gun_level") == "L"),
                "M": sum(int(lot.get("qty_available", 0)) for lot in lots if lot.get("gun_level") == "M"),
                "H": sum(int(lot.get("qty_available", 0)) for lot in lots if lot.get("gun_level") == "H"),
            }

        return lots, level_totals

    async def _save_state(self, state: dict[str, Any]) -> bool:
        wholesale_file = Path(getattr(self, "wholesale_inventory_file", self.state_file))
        store_file = Path(getattr(self, "store_state_file", self.state_file))
        default_data_dir = Path(getattr(self, "data_dir", Path(self.state_file).parent))
        store_inventory_dir = Path(getattr(self, "store_inventory_dir", default_data_dir / "inventory" / "stores"))

        # Keep a compatibility snapshot in the main state file so deployments
        # that only persist `state.json` (and not the split inventory files)
        # still retain stock across restarts.
        state_main = dict(state)

        wholesale_payload = {
            "wholesale_lots": state.get("wholesale_lots", []),
            "updated_at": self._now_iso(),
        }

        stores_payload = {
            "shop_registry": state.get("shop_registry", {}),
            "stores": {},
            "updated_at": self._now_iso(),
        }

        main_ok = await helpers.save_json_file(self.state_file, state_main)
        wholesale_ok = await helpers.save_json_file(wholesale_file, wholesale_payload)

        store_index_ok = True

        store_inventory_dir.mkdir(parents=True, exist_ok=True)
        stores = state.get("stores", {})
        desired_paths: set[Path] = set()
        store_files_ok = True
        if isinstance(stores, dict):
            for store_id, store_data in stores.items():
                dest = self._store_inventory_file(str(store_id))
                desired_paths.add(dest)
                stores_payload["stores"][store_id] = {
                    "owner_id": store_data.get("owner_id"),
                    "inventory_file": str(dest),
                }
                payload = {
                    "store_id": store_id,
                    "owner_id": store_data.get("owner_id"),
                    "lots": store_data.get("lots", []),
                    "updated_at": self._now_iso(),
                }
                if not await helpers.save_json_file(dest, payload):
                    store_files_ok = False

        for old_file in self._list_store_inventory_files():
            if old_file not in desired_paths:
                try:
                    old_file.unlink(missing_ok=True)
                except Exception:
                    logger.exception("Failed to remove stale store inventory file: %s", old_file)
                    store_files_ok = False

        if store_file != self.state_file:
            store_index_ok = await helpers.save_json_file(store_file, stores_payload)

        ok = main_ok and wholesale_ok and store_index_ok and store_files_ok
        if not ok:
            logger.error(
                "Wholesaler persistence failure main_ok=%s wholesale_ok=%s store_index_ok=%s store_files_ok=%s "
                "state_file=%s wholesale_file=%s store_file=%s store_inventory_dir=%s",
                main_ok,
                wholesale_ok,
                store_index_ok,
                store_files_ok,
                self.state_file,
                wholesale_file,
                store_file,
                store_inventory_dir,
            )
        return ok

    async def _append_tx(self, tx: dict[str, Any]) -> bool:
        return await helpers.append_json_file(self.tx_file, tx)

    @staticmethod
    def _coerce_role_ids(value: Any) -> set[int]:
        """Normalize role-id config values from int/list/CSV formats."""
        if value is None:
            return set()
        if isinstance(value, (int, float)):
            return {int(value)}
        if isinstance(value, str):
            parts = [part.strip() for part in value.split(",")]
            return {
                int(part)
                for part in parts
                if part and part.lstrip("-").isdigit()
            }

        role_ids: set[int] = set()
        try:
            iterator = iter(value)
        except TypeError:
            return set()

        for item in iterator:
            if isinstance(item, (int, float)):
                role_ids.add(int(item))
                continue
            text = str(item).strip()
            if text and text.lstrip("-").isdigit():
                role_ids.add(int(text))
        return role_ids

    def _is_admin(self, member: discord.Member) -> bool:
        admin_role_ids = self._coerce_role_ids(getattr(config, "WHOLESALER_ADMIN_ROLE_IDS", []))
        return bool(member.guild_permissions.administrator) or any(r.id in admin_role_ids for r in member.roles)

    def _is_store_owner(self, member: discord.Member) -> bool:
        store_role_ids = self._coerce_role_ids(getattr(config, "WHOLESALER_STORE_ROLE_IDS", []))
        return any(r.id in store_role_ids for r in member.roles)

    async def _audit_send(self, text: str) -> None:
        channel_id = int(getattr(config, "WHOLESALER_AUDIT_CHANNEL_ID", 0) or 0)
        if channel_id <= 0:
            logger.warning("Missing wholesaler audit channel id=%s", channel_id)
            return

        channel = self.bot.get_channel(channel_id)
        if channel is None:
            try:
                channel = await self.bot.fetch_channel(channel_id)
            except Exception:
                logger.warning("Missing wholesaler audit channel id=%s", channel_id)
                return

        if not channel:
            logger.warning("Missing wholesaler audit channel id=%s", channel_id)
            return
        try:
            await channel.send(text)
        except Exception:
            logger.exception("Failed to send wholesaler audit line")


    async def _system_enabled(self, ctx: commands.Context) -> bool:
        control = self.bot.get_cog("SystemControl")
        if control and not control.is_enabled("wholesaler"):
            await ctx.send("⚠️ The wholesaler system is currently disabled.")
            return False
        return True

    async def _ensure_member(self, ctx: commands.Context) -> Optional[discord.Member]:
        if not ctx.guild or not isinstance(ctx.author, discord.Member):
            await ctx.send("❌ This command can only be used in the server.")
            return None
        return ctx.author

    @staticmethod
    def _store_id(guild_id: int, owner_id: int) -> str:
        return f"{guild_id}:{owner_id}"

    def _shop_display_name(
        self,
        state: dict[str, Any],
        owner_id: int,
        requested_shop: Optional[str] = None,
    ) -> str:
        if requested_shop:
            return self._normalize_shop_name(requested_shop)

        registry = state.get("shop_registry", {})
        aliases = sorted(name for name, mapped_owner in registry.items() if int(mapped_owner) == int(owner_id))
        if aliases:
            return aliases[0]

        return f"owner:{owner_id}"

    def _build_tx(self, tx_type: str, **kwargs: Any) -> dict[str, Any]:
        return {
            "tx_id": f"tx-{uuid.uuid4().hex[:12]}",
            "type": tx_type,
            "timestamp": self._now_iso(),
            "status": "SUCCESS",
            "error_details": "",
            **kwargs,
        }

    async def _resolve_store_owner_id(
        self, ctx: commands.Context, state: dict[str, Any], shop_or_mention: Optional[str], default_owner: int
    ) -> int:
        if not shop_or_mention:
            return default_owner
        if shop_or_mention.startswith("<@") and shop_or_mention.endswith(">"):
            return int(shop_or_mention.strip("<@!>"))
        key = self._normalize_shop_name(shop_or_mention)
        return int(state.get("shop_registry", {}).get(key, default_owner))

    async def _get_total_balance(self, user_id: int) -> Optional[tuple[int, int, int]]:
        balance = await self.unbelievaboat.get_balance(user_id)
        if not balance:
            return None
        cash = int(balance.get("cash", 0))
        bank = int(balance.get("bank", 0))
        return cash, bank, cash + bank

    async def _deduct_funds(self, user_id: int, amount: int, reason: str) -> tuple[bool, str]:
        balances = await self._get_total_balance(user_id)
        if balances is None:
            return False, "Unable to fetch user balance"
        cash, _bank, total = balances
        if total < amount:
            return False, f"Insufficient funds (${total}/${amount})"

        cash_deduct = min(max(cash, 0), amount)
        bank_deduct = max(0, amount - cash_deduct)
        payload: dict[str, int] = {}
        if cash_deduct:
            payload["cash"] = -cash_deduct
        if bank_deduct:
            payload["bank"] = -bank_deduct
        ok = await self.unbelievaboat.update_balance(user_id, payload, reason=reason)
        return (ok, "" if ok else "Failed to deduct funds")

    async def _credit_funds(self, user_id: int, amount: int, reason: str) -> bool:
        return await self.unbelievaboat.update_balance(user_id, {"cash": amount}, reason=reason)

    @commands.command(name="wh_setshop")
    async def wh_setshop(self, ctx: commands.Context, shop_name: str, owner: discord.Member):
        """Bind shop aliases (shop1/shop2/shop3 etc.) to an owner Discord account."""
        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        if not self._is_admin(member):
            await ctx.send("❌ Admin role required.")
            return

        normalized = self._normalize_shop_name(shop_name)
        if not normalized:
            await ctx.send("❌ Invalid shop name.")
            return

        async with self.lock:
            state = await self._load_state()
            state.setdefault("shop_registry", {})[normalized] = owner.id
            await self._save_state(state)

        await ctx.send(f"✅ `{normalized}` is now mapped to {owner.mention}.")
        await self._audit_send(
            f"[WHOLESALE_SET_SHOP] by={member.mention} shop={normalized} owner={owner.mention}"
        )

    @commands.command(name="wh_shops")
    async def wh_shops(self, ctx: commands.Context):

        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        state = await self._load_state()
        registry = state.get("shop_registry", {})
        if not registry:
            await ctx.send("No shop aliases configured.")
            return
        lines = [f"`{name}` → <@{owner_id}>" for name, owner_id in sorted(registry.items())]
        await ctx.send("**Shop Registry**\n" + "\n".join(lines[:30]))

    @commands.command(name="wh_list")
    async def wh_list(self, ctx: commands.Context):

        if not await self._system_enabled(ctx):
            return
        state = await self._load_state()
        lots = [lot for lot in state.get("wholesale_lots", []) if int(lot.get("qty_available", 0)) > 0]
        if not lots:
            await ctx.send("No wholesale lots available.")
            return
        lines = [
            f"`{l['lot_id']}` | {l['gun_name']} ({l['gun_level']}) | ${l['unit_cost']} | qty {l['qty_available']}"
            for l in lots[:25]
        ]
        await ctx.send("**Wholesaler Stock**\n" + "\n".join(lines))

    @commands.command(name="store_inv")
    async def store_inv(self, ctx: commands.Context, *, shop: Optional[str] = None):
        """Show your inventory or a named shop inventory (`!store_inv shop1`)."""
        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return

        state = await self._load_state()
        owner_id = await self._resolve_store_owner_id(ctx, state, shop, member.id)
        if owner_id != member.id and not self._is_admin(member):
            await ctx.send("❌ Only admins can inspect other shops.")
            return

        store_id = self._store_id(ctx.guild.id, owner_id)
        lots = [l for l in state.get("stores", {}).get(store_id, {}).get("lots", []) if l.get("qty_remaining", 0) > 0]
        if not lots:
            await ctx.send("Store inventory is empty.")
            return

        shop_title = self._shop_display_name(state, owner_id, shop)
        lines = [
            f"`{l['lot_id']}` | {l['gun_name']} ({l['gun_level']}) | cost ${l['unit_cost']} | qty {l['qty_remaining']}"
            for l in lots[:30]
        ]
        await ctx.send(f"**Store Inventory ({shop_title})**\n" + "\n".join(lines))

    @commands.command(name="wh_buy")
    async def wh_buy(self, ctx: commands.Context, lot_id: str, qty: int):

        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        if not self._is_store_owner(member):
            await ctx.send("❌ Store owner role required.")
            return
        if qty <= 0:
            await ctx.send("❌ qty must be > 0.")
            return

        async with self.lock:
            state = await self._load_state()
            lot = next((l for l in state.get("wholesale_lots", []) if l.get("lot_id") == lot_id), None)
            if not lot or int(lot.get("qty_available", 0)) < qty:
                await ctx.send("❌ Lot unavailable or insufficient quantity.")
                return

            total = int(lot["unit_cost"]) * qty
            ok, err = await self._deduct_funds(member.id, total, f"Wholesale purchase {lot_id}")
            tx = self._build_tx(
                "WHOLESALE_BUY",
                seller_id="WHOLESALER",
                buyer_id=member.id,
                gun_name=lot["gun_name"],
                gun_level=lot["gun_level"],
                qty=qty,
                unit_price=lot["unit_cost"],
                total_price=total,
                lot_id=lot_id,
            )
            if not ok:
                tx["status"] = "FAILED"
                tx["error_details"] = err
                await self._append_tx(tx)
                await ctx.send(f"❌ Purchase failed: {err}")
                return

            lot["qty_available"] -= qty
            store_id = self._store_id(ctx.guild.id, member.id)
            store = state.setdefault("stores", {}).setdefault(store_id, {"owner_id": member.id, "lots": []})
            if not isinstance(store, dict):
                store = {"owner_id": member.id, "lots": []}
                state["stores"][store_id] = store
            lots = store.get("lots", [])
            if not isinstance(lots, list):
                lots = []
                store["lots"] = lots
            existing = next((l for l in lots if isinstance(l, dict) and l.get("lot_id") == lot_id), None)
            if existing:
                existing["qty_remaining"] += qty
            else:
                lots.append(
                    {
                        "lot_id": lot_id,
                        "gun_name": lot["gun_name"],
                        "gun_level": lot["gun_level"],
                        "unit_cost": lot["unit_cost"],
                        "qty_remaining": qty,
                    }
                )
            await self._save_state(state)
            await self._append_tx(tx)

        await ctx.send(f"✅ Purchased {qty}x {lot['gun_name']} for ${total}.")
        await self._audit_send(
            f"[WHOLESALE_BUY] tx={tx['tx_id']} buyer={member.mention} gun={lot['gun_name']} level={lot['gun_level']} qty={qty} total={total} lot={lot_id}"
        )

    @commands.command(name="sell")
    async def sell(
        self,
        ctx: commands.Context,
        buyer: discord.Member,
        lot_id: str,
        qty: int,
        total_price: int,
        *,
        extra: str = "",
    ):
        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        if not self._is_store_owner(member):
            await ctx.send("❌ Store owner role required.")
            return
        if qty <= 0 or total_price <= 0:
            await ctx.send("❌ qty and total_price must be > 0.")
            return

        character_name = ""
        if "character:" in extra.lower():
            character_name = extra.split(":", 1)[1].strip().strip('"')

        async with self.lock:
            state = await self._load_state()
            store_id = self._store_id(ctx.guild.id, member.id)
            store = state.get("stores", {}).get(store_id)
            if not store:
                await ctx.send("❌ No store inventory found.")
                return

            store_lot = next((l for l in store.get("lots", []) if l.get("lot_id") == lot_id), None)
            if not store_lot or int(store_lot.get("qty_remaining", 0)) < qty:
                await ctx.send("❌ Invalid lot or insufficient quantity.")
                return

            tx = self._build_tx(
                "PLAYER_SALE",
                seller_id=member.id,
                buyer_id=buyer.id,
                gun_name=store_lot["gun_name"],
                gun_level=store_lot["gun_level"],
                qty=qty,
                unit_price=max(1, total_price // qty),
                total_price=total_price,
                lot_id=lot_id,
                character_name=character_name,
            )

            deduct_ok, deduct_err = await self._deduct_funds(
                buyer.id,
                total_price,
                f"Gun purchase from {member.id} ({store_lot['gun_name']})",
            )
            if not deduct_ok:
                tx["status"] = "FAILED"
                tx["error_details"] = deduct_err
                await self._append_tx(tx)
                await ctx.send(f"❌ Sale failed: {deduct_err}")
                return

            payout_ok = await self._credit_funds(
                member.id,
                total_price,
                f"Gun sale to {buyer.id} ({store_lot['gun_name']})",
            )
            if not payout_ok:
                tx["status"] = "PENDING_PAYOUT"
                tx["error_details"] = "Buyer charged, seller payout failed"
                state.setdefault("pending_payouts", []).append(
                    {"tx_id": tx["tx_id"], "seller_id": member.id, "amount": total_price}
                )
                await self._save_state(state)
                await self._append_tx(tx)
                await self._audit_send(
                    f"🚨 [PENDING_PAYOUT] tx={tx['tx_id']} seller={member.mention} buyer={buyer.mention} amount={total_price}"
                )
                await ctx.send("⚠️ Buyer charged, seller payout pending admin retry.")
                return

            store_lot["qty_remaining"] -= qty
            await self._save_state(state)
            await self._append_tx(tx)

        await ctx.send(f"✅ Sold {qty}x {store_lot['gun_name']} for ${total_price}.")
        await self._audit_send(
            "[PLAYER_SALE_RECEIPT] "
            f"tx={tx['tx_id']} ts={tx['timestamp']} seller={member.mention} buyer={buyer.mention} "
            f"character={character_name or 'N/A'} gun={store_lot['gun_name']} level={store_lot['gun_level']} "
            f"qty={qty} total={total_price} unit_cost={store_lot['unit_cost']} lot={lot_id}"
        )

    @commands.command(name="wh_restock")
    async def wh_restock(self, ctx: commands.Context, seed: Optional[int] = None):

        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        if not self._is_admin(member):
            await ctx.send("❌ Admin role required.")
            return

        try:
            guns = await self._load_master_guns()
        except Exception as e:
            logger.exception("wh_restock failed")
            await ctx.send(f"❌ Restock failed while reading source sheet: {e}")
            return

        if not guns:
            await ctx.send("❌ No valid guns found in source sheet.")
            return

        async with self.lock:
            state = await self._load_state()
            cfg = self._resolve_restock_settings(state)
            rng = random.Random(seed)
            lots, level_totals = self._generate_restock_lots(guns, cfg, rng)

            state["wholesale_lots"] = lots
            state.setdefault("settings", {}).setdefault("restock", {}).update(cfg)
            saved = await self._save_state(state)

        if not saved:
            logger.error(
                "wh_restock: _save_state returned False — files may not have been written "
                "(state=%s, wholesale=%s, store=%s)",
                self.state_file,
                self.wholesale_inventory_file,
                self.store_state_file,
            )
            await ctx.send(
                f"⚠️ Restocked {len(lots)} lots but **failed to persist** inventory files. "
                "Check bot logs for details."
            )
            return

        await ctx.send(f"✅ Wholesaler is restocked. Added {len(lots)} wholesale lots.")
        await self._audit_send(
            f"[WHOLESALE_RESTOCK] by={member.mention} lots={len(lots)} qtyL={level_totals['L']} qtyM={level_totals['M']} qtyH={level_totals['H']}"
        )

    @commands.command(name="wh_clear_inventory")
    async def wh_clear_inventory(self, ctx: commands.Context):
        """Clear all current wholesaler lots while preserving store inventories."""
        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        if not self._is_admin(member):
            await ctx.send("❌ Admin role required.")
            return

        async with self.lock:
            state = await self._load_state()
            lots = state.get("wholesale_lots", [])
            lot_count = len(lots)
            qty_count = sum(max(int(lot.get("qty_available", 0)), 0) for lot in lots)
            state["wholesale_lots"] = []
            await self._save_state(state)

        await ctx.send(f"✅ Cleared wholesaler inventory ({lot_count} lots, {qty_count} units).")
        await self._audit_send(
            f"[WHOLESALE_CLEAR_INVENTORY] by={member.mention} lots={lot_count} units={qty_count}"
        )

    async def auto_refresh_weekly_after_cyberware(self) -> bool:
        """Auto-restock once per week, called by cyberware weekly process."""
        return await self._auto_restock_if_due(datetime.now(timezone.utc), trigger="CYBERWARE")

    async def _auto_restock_if_due(self, now: datetime, trigger: str) -> bool:
        """Run one weekly restock if we have not yet refreshed this Sunday."""
        control = self.bot.get_cog("SystemControl")
        if control and not control.is_enabled("wholesaler"):
            return False
        try:
            guns = await self._load_master_guns()
        except Exception:
            logger.exception("Auto wholesaler refresh failed during sheet read")
            return False

        if not guns:
            return False

        sunday_key = now.strftime("%Y-%m-%d")
        async with self.lock:
            state = await self._load_state()
            settings = state.setdefault("settings", {})
            last_key = str(settings.get("last_auto_restock_sunday", ""))
            if last_key == sunday_key:
                return True

            # Legacy key fallback to avoid duplicate restock in the same week.
            legacy_week_key = now.strftime("%Y-W%U")
            if str(settings.get("last_auto_restock_week", "")) == legacy_week_key:
                settings["last_auto_restock_sunday"] = sunday_key
                await self._save_state(state)
                return True

            cfg = self._resolve_restock_settings(state)
            rng = random.Random()
            lots, _level_totals = self._generate_restock_lots(guns, cfg, rng)

            state["wholesale_lots"] = lots
            settings["last_auto_restock_sunday"] = sunday_key
            settings["last_auto_restock_week"] = legacy_week_key
            saved = await self._save_state(state)
            if not saved:
                logger.error("Auto restock save failed trigger=%s", trigger)

        await self._audit_send(
            f"[WHOLESALE_AUTO_RESTOCK] trigger={trigger} lots={len(lots)} sunday={sunday_key}"
        )
        return True

    @commands.command(name="wh_restock_settings")
    async def wh_restock_settings(
        self,
        ctx: commands.Context,
        key: Optional[str] = None,
        value: Optional[int] = None,
    ):
        """View or update weekly restock settings.

        Example: !wh_restock_settings lots_L 12
        Type mix example: !wh_restock_settings revolver_L 3
        """
        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        if not self._is_admin(member):
            await ctx.send("❌ Admin role required.")
            return

        async with self.lock:
            state = await self._load_state()
            cfg = self._resolve_restock_settings(state)

            if key and value is not None:
                type_keys = {f"{weapon_type}_{lvl}" for weapon_type in self.WEAPON_TYPES for lvl in ("L", "M", "H")}
                if key not in self.DEFAULT_RESTOCK_SETTINGS and key not in type_keys:
                    await ctx.send(
                        "❌ Invalid key. Use one of: "
                        + ", ".join(sorted(list(self.DEFAULT_RESTOCK_SETTINGS.keys()) + list(type_keys)))
                    )
                    return
                if key in type_keys:
                    cfg[key] = max(0, int(value))
                    restock_type_lots = state.setdefault("settings", {}).setdefault("restock_type_lots", {})
                    restock_type_lots[key] = cfg[key]
                elif key in {"lots_L", "lots_M", "lots_H"}:
                    cfg[key] = max(0, int(value))
                    state.setdefault("settings", {}).setdefault("restock", {})[key] = cfg[key]
                else:
                    cfg[key] = max(1, int(value))
                    state.setdefault("settings", {}).setdefault("restock", {})[key] = cfg[key]
                await self._save_state(state)
                await ctx.send(f"✅ Updated {key} to {cfg[key]}.")
                await self._audit_send(
                    f"[WHOLESALE_RESTOCK_SETTINGS] by={member.mention} key={key} value={cfg[key]}"
                )
                return

        lines = ["**Wholesaler Restock Settings**"]
        for k in sorted(self.DEFAULT_RESTOCK_SETTINGS.keys()):
            lines.append(f"`{k}` = {cfg[k]}")
        lines.append("\n**Type + Size Lot Targets**")
        lines.append("(set any value > 0 to use explicit type mix during restock)")
        for weapon_type in self.WEAPON_TYPES:
            pretty = weapon_type.replace("_", " ").title()
            lines.append(
                f"`{weapon_type}_L`={cfg[f'{weapon_type}_L']} | "
                f"`{weapon_type}_M`={cfg[f'{weapon_type}_M']} | "
                f"`{weapon_type}_H`={cfg[f'{weapon_type}_H']}  ({pretty})"
            )
        await ctx.send("\n".join(lines))

    @commands.command(name="wh_setsheet")
    async def wh_setsheet(self, ctx: commands.Context, *, xlsx_export_url: str):
        """Set/clear runtime Google Sheets XLSX export URL for wholesaler source.

        Use `!wh_setsheet off` to clear runtime override and fall back to config.
        """
        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        if not self._is_admin(member):
            await ctx.send("❌ Admin role required.")
            return

        value = self._clean_sheet_url_input(xlsx_export_url)
        async with self.lock:
            state = await self._load_state()
            settings = state.setdefault("settings", {})
            if value.lower() in {"off", "none", "clear"}:
                settings.pop("master_sheet_url", None)
                await self._save_state(state)
                await ctx.send("✅ Runtime sheet URL cleared. Using config/default source.")
                return

            if not value.startswith(("http://", "https://")):
                await ctx.send("❌ URL must start with http/https.")
                return

            normalized = self._normalize_sheet_source_url(value)
            settings["master_sheet_url"] = normalized
            await self._save_state(state)

        await ctx.send("✅ Runtime wholesaler sheet URL updated.")
        await self._audit_send(f"[WHOLESALE_SOURCE_SET] by={member.mention} url={normalized}")

    @commands.command(name="wh_recheck")
    async def wh_recheck(self, ctx: commands.Context):
        """Reconcile current wholesaler lots against current sheet prices/levels."""
        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        if not self._is_admin(member):
            await ctx.send("❌ Admin role required.")
            return

        try:
            guns = await self._load_master_guns()
        except Exception as e:
            await ctx.send(f"❌ Recheck failed: {e}")
            return

        index = {g["gun_name"].strip().lower(): g for g in guns}
        state = await self._load_state()
        missing = []
        price_mismatch = []
        level_mismatch = []
        for lot in state.get("wholesale_lots", []):
            g = index.get(str(lot.get("gun_name", "")).strip().lower())
            if not g:
                missing.append(lot)
                continue
            if int(g["price_new"]) != int(lot.get("unit_cost", 0)):
                price_mismatch.append((lot, g["price_new"]))
            if g["gun_level"] != lot.get("gun_level"):
                level_mismatch.append((lot, g["gun_level"]))

        lines = [
            f"Checked {len(state.get('wholesale_lots', []))} lots against {len(guns)} sheet rows.",
            f"Missing in sheet: {len(missing)}",
            f"Price mismatches: {len(price_mismatch)}",
            f"Level mismatches: {len(level_mismatch)}",
        ]
        if missing[:3]:
            lines.append("Missing examples: " + ", ".join(m["gun_name"] for m in missing[:3]))
        if price_mismatch[:3]:
            lines.append(
                "Price examples: " + ", ".join(f"{x[0]['gun_name']} lot=${x[0]['unit_cost']} sheet=${x[1]}" for x in price_mismatch[:3])
            )
        if level_mismatch[:3]:
            lines.append(
                "Level examples: " + ", ".join(f"{x[0]['gun_name']} lot={x[0]['gun_level']} sheet={x[1]}" for x in level_mismatch[:3])
            )

        await ctx.send("\n".join(lines))
        await self._audit_send("[WHOLESALE_RECHECK] " + " | ".join(lines[:4]))

    @commands.command(name="wh_paths")
    async def wh_paths(self, ctx: commands.Context):
        """Show resolved wholesaler persistence paths and whether files exist."""
        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        if not self._is_admin(member):
            await ctx.send("❌ Admin role required.")
            return

        files = (
            ("data_dir", Path(self.data_dir), True),
            ("state_file", Path(self.state_file), False),
            ("store_state_file", Path(self.store_state_file), False),
            ("wholesale_inventory_file", Path(self.wholesale_inventory_file), False),
            ("store_inventory_dir", Path(self.store_inventory_dir), True),
            ("transactions_file", Path(self.tx_file), False),
        )
        lines = ["**Wholesaler persistence paths**"]
        for label, path, is_dir in files:
            exists = path.is_dir() if is_dir else path.is_file()
            lines.append(f"- `{label}`: `{path}` (exists={exists})")

        configured_dir = getattr(config, "WHOLESALER_DATA_DIR", None)
        if configured_dir:
            lines.append(f"- `WHOLESALER_DATA_DIR` config: `{configured_dir}`")

        await ctx.send("\n".join(lines))

    @commands.command(name="wh_add")
    async def wh_add(self, ctx: commands.Context, gun_name: str, level: str, unit_cost: int, qty: int):

        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        if not self._is_admin(member):
            await ctx.send("❌ Admin role required.")
            return
        if unit_cost <= 0 or qty <= 0:
            await ctx.send("❌ unit_cost and qty must be positive.")
            return

        level = level.upper()
        if level not in {"L", "M", "H"}:
            await ctx.send("❌ level must be L/M/H.")
            return

        lot = {
            "lot_id": f"lot-{datetime.now(timezone.utc).strftime('%Y%m%d')}-{uuid.uuid4().hex[:6]}",
            "gun_name": gun_name,
            "gun_level": level,
            "unit_cost": unit_cost,
            "qty_available": qty,
            "created_at": self._now_iso(),
        }
        async with self.lock:
            state = await self._load_state()
            state.setdefault("wholesale_lots", []).append(lot)
            await self._save_state(state)

        tx = self._build_tx(
            "ADMIN_ADJUST",
            seller_id=member.id,
            buyer_id="WHOLESALER",
            gun_name=gun_name,
            gun_level=level,
            qty=qty,
            unit_price=unit_cost,
            total_price=unit_cost * qty,
            lot_id=lot["lot_id"],
        )
        await self._append_tx(tx)
        await ctx.send(f"✅ Added lot `{lot['lot_id']}`.")
        await self._audit_send(
            f"[WHOLESALE_ADMIN_ADD] by={member.mention} gun={gun_name} level={level} qty={qty} unit_cost={unit_cost} lot={lot['lot_id']}"
        )

    @commands.command(name="store_add")
    async def store_add(
        self,
        ctx: commands.Context,
        store_owner: discord.Member,
        gun_name: str,
        level: str,
        unit_cost: int,
        qty: int,
    ):
        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        if not self._is_admin(member):
            await ctx.send("❌ Admin role required.")
            return
        if unit_cost <= 0 or qty <= 0:
            await ctx.send("❌ unit_cost and qty must be positive.")
            return

        level = level.upper()
        if level not in {"L", "M", "H"}:
            await ctx.send("❌ level must be L/M/H.")
            return

        lot_id = f"lot-{datetime.now(timezone.utc).strftime('%Y%m%d')}-{uuid.uuid4().hex[:6]}"
        store_lot = {
            "lot_id": lot_id,
            "gun_name": gun_name,
            "gun_level": level,
            "unit_cost": unit_cost,
            "qty_remaining": qty,
        }
        async with self.lock:
            state = await self._load_state()
            store_id = self._store_id(ctx.guild.id, store_owner.id)
            store = state.setdefault("stores", {}).setdefault(store_id, {"owner_id": store_owner.id, "lots": []})
            store["lots"].append(store_lot)
            await self._save_state(state)

        tx = self._build_tx(
            "ADMIN_ADJUST",
            seller_id=member.id,
            buyer_id=store_owner.id,
            gun_name=gun_name,
            gun_level=level,
            qty=qty,
            unit_price=unit_cost,
            total_price=unit_cost * qty,
            lot_id=lot_id,
        )
        await self._append_tx(tx)
        await ctx.send(f"✅ Added `{gun_name}` to {store_owner.mention} inventory (lot `{lot_id}`).")
        await self._audit_send(
            f"[STORE_ADMIN_ADD] by={member.mention} owner={store_owner.mention} gun={gun_name} level={level} qty={qty} unit_cost={unit_cost} lot={lot_id}"
        )

    @commands.command(name="wh_tx")
    async def wh_tx(self, ctx: commands.Context, tx_id: str):

        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        if not self._is_admin(member):
            await ctx.send("❌ Admin role required.")
            return

        txs = await helpers.load_json_file(self.tx_file, default=[])
        tx = next((t for t in txs if t.get("tx_id") == tx_id), None)
        if not tx:
            await ctx.send("Transaction not found.")
            return
        await ctx.send(f"```json\n{tx}\n```")

    @commands.command(name="wh_retry_payout")
    async def wh_retry_payout(self, ctx: commands.Context, tx_id: str):

        if not await self._system_enabled(ctx):
            return
        member = await self._ensure_member(ctx)
        if not member:
            return
        if not self._is_admin(member):
            await ctx.send("❌ Admin role required.")
            return

        async with self.lock:
            state = await self._load_state()
            pending = state.get("pending_payouts", [])
            entry = next((p for p in pending if p.get("tx_id") == tx_id), None)
            if not entry:
                await ctx.send("No pending payout found.")
                return

            ok = await self._credit_funds(int(entry["seller_id"]), int(entry["amount"]), f"Retry payout {tx_id}")
            if not ok:
                await ctx.send("❌ Retry failed.")
                return

            state["pending_payouts"] = [p for p in pending if p.get("tx_id") != tx_id]
            await self._save_state(state)

        tx = self._build_tx(
            "REFUND",
            seller_id="SYSTEM",
            buyer_id=entry["seller_id"],
            gun_name="PAYOUT_RETRY",
            gun_level="N/A",
            qty=1,
            unit_price=entry["amount"],
            total_price=entry["amount"],
            lot_id=tx_id,
        )
        await self._append_tx(tx)
        await self._audit_send(
            f"[PAYOUT_RETRY_SUCCESS] tx={tx_id} seller=<@{entry['seller_id']}> amount={entry['amount']}"
        )
        await ctx.send("✅ Payout retried.")
