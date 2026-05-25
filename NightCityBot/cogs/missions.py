"""Mission tracking cog.

Fixer-only commands for tracking which players have been on missions
and how often, so the GM team can keep mission invites fair.

Commands:
  !mission_check <users...>
      Report most-recent mission date, days since, and total count
      for each user.

  !mission_record <users...> [date=YYYY-MM-DD | MM/DD/YYYY]
      Record a mission for each user. Date defaults to today (UTC).

``parse_mission_sheet`` is retained as a module-level helper for any
future one-shot roster imports, but no Discord command exposes it now —
the original roster sheet was imported directly via a dev script.
"""
from __future__ import annotations

import logging
import re
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

import discord
from discord.ext import commands, tasks
from openpyxl import load_workbook

from NightCityBot.utils.db import (
    mission_log_get,
    mission_log_record,
    mission_event_get,
    mission_event_get_for_user,
    mission_event_list_due,
    mission_event_mark_paid,
    mission_event_claim_for_payout,
    mission_event_unclaim,
    mission_event_update,
    mission_event_cancel,
)
from NightCityBot.utils.permissions import is_fixer


logger = logging.getLogger(__name__)


async def _post_mission_audit(bot, **kwargs) -> None:
    """Lazy wrapper to avoid a circular import with fixer_hub."""
    try:
        from NightCityBot.cogs.fixer_hub import post_mission_audit as _impl
    except Exception:
        logger.error("Could not import post_mission_audit", exc_info=True)
        return
    await _impl(bot, **kwargs)

# Only events whose title begins with this prefix (case-insensitive) are
# treated as bot-managed missions during reconciliation. Anything else
# (Main Session, Social, etc.) is ignored / treated as canceled.
ACTORS_NEEDED_PREFIX = "Actors Needed:"


def _strip_actors_prefix(title: str) -> str:
    """Return the mission name part after the 'Actors Needed:' prefix."""
    t = (title or "").strip()
    if t.lower().startswith(ACTORS_NEEDED_PREFIX.lower()):
        return t[len(ACTORS_NEEDED_PREFIX):].strip()
    return t


def _dt_close(a: Optional[datetime], b: Optional[datetime], tol_seconds: int = 30) -> bool:
    """Treat two datetimes as equal if within `tol_seconds` of each other.

    Discord round-trips timestamps as ISO strings, so allow a small tolerance.
    """
    if a is None or b is None:
        return a is b
    if a.tzinfo is None:
        a = a.replace(tzinfo=timezone.utc)
    if b.tzinfo is None:
        b = b.replace(tzinfo=timezone.utc)
    return abs((a - b).total_seconds()) <= tol_seconds

# Auto-payout fires at midnight US Eastern on the day AFTER the mission
# starts (per user spec). Using America/New_York handles EST/EDT automatically.
PAYOUT_TZ = ZoneInfo("America/New_York")


def compute_payout_ts(after_utc: datetime) -> datetime:
    """Return the UTC timestamp of the next 00:00 America/New_York strictly
    after ``after_utc``.

    Callers should pass the mission's **end** time so payout never fires
    before the mission has finished (a mission spanning midnight ET would
    otherwise be paid mid-session).

    Example: end = 2026-05-22 20:00 UTC (16:00 ET on May 22) →
    payout = 2026-05-23 00:00 ET = 2026-05-23 04:00 UTC.
    """
    if after_utc.tzinfo is None:
        after_utc = after_utc.replace(tzinfo=timezone.utc)
    after_et = after_utc.astimezone(PAYOUT_TZ)
    next_day_et = (after_et + timedelta(days=1)).date()
    midnight_et = datetime(
        next_day_et.year, next_day_et.month, next_day_et.day,
        0, 0, 0, tzinfo=PAYOUT_TZ,
    )
    return midnight_et.astimezone(timezone.utc)

_MENTION_RE = re.compile(r"^<@!?(\d+)>$")
_ID_RE = re.compile(r"^\d{15,22}$")
_DATE_SLASH_RE = re.compile(r"^\s*(\d{1,2})/(\d{1,2})/(\d{2,4})\s*$")
_DATE_DASH_RE = re.compile(r"^\s*(\d{4})-(\d{1,2})-(\d{1,2})\s*$")


def _parse_date_token(token: str) -> Optional[date]:
    """Parse a YYYY-MM-DD or M/D/YY(YY) date token."""
    if not token:
        return None
    m = _DATE_DASH_RE.match(token)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = _DATE_SLASH_RE.match(token)
    if m:
        mo, da, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yr < 100:
            yr += 2000
        try:
            return date(yr, mo, da)
        except ValueError:
            return None
    return None


def _parse_sheet_dates(cell: object) -> list[date]:
    """Parse the comma-separated date list from sheet column D."""
    if cell is None:
        return []
    if isinstance(cell, datetime):
        return [cell.date()]
    if isinstance(cell, date):
        return [cell]
    text = str(cell).strip()
    if not text:
        return []
    out: list[date] = []
    for raw in text.split(","):
        d = _parse_date_token(raw)
        if d is not None:
            out.append(d)
    return out


def parse_mission_sheet(xlsx_path: Path | str) -> list[dict]:
    """Parse the mission roster XLSX into normalized rows.

    Returns a list of dicts with keys ``user_id``, ``username``,
    ``mission_count``, ``mission_dates``. Skips rows whose column A
    doesn't look like a Discord snowflake.
    """
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        out: list[dict] = []
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            a = row[0]
            if a is None:
                continue
            raw_id = str(a).strip()
            if not raw_id.isdigit() or not (15 <= len(raw_id) <= 22):
                # Header rows or junk rows.
                continue
            username = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            try:
                count = int(row[2]) if len(row) > 2 and row[2] not in (None, "") else 0
            except (TypeError, ValueError):
                count = 0
            dates = _parse_sheet_dates(row[3]) if len(row) > 3 else []
            # If count is missing/zero but we parsed dates, infer count.
            if count <= 0 and dates:
                count = len(dates)
            out.append({
                "user_id": raw_id,
                "username": username,
                "mission_count": count,
                "mission_dates": dates,
            })
        return out
    finally:
        wb.close()


async def _resolve_target(ctx: commands.Context, token: str) -> tuple[str, str, Optional[discord.abc.User]]:
    """Resolve one CLI token into (user_id, username, user_obj_or_None).

    Accepts:
      - Discord mention (<@123> or <@!123>)
      - Raw snowflake id (digits only)
      - Plain username / display name (searched against the guild)
    Falls back to (token, token, None) when nothing matches so the caller
    can still record / look up by raw id.
    """
    token = token.strip()
    if not token:
        return ("", "", None)

    async def _fetch(uid_str: str):
        member = ctx.guild.get_member(int(uid_str)) if ctx.guild else None
        if member is None and ctx.bot is not None:
            try:
                member = await ctx.bot.fetch_user(int(uid_str))
            except Exception:
                member = None
        return member

    # Mention?
    m = _MENTION_RE.match(token)
    if m:
        uid = m.group(1)
        member = await _fetch(uid)
        uname = getattr(member, "display_name", None) or getattr(member, "name", "") or uid
        return (uid, uname, member)

    # Raw id?
    if _ID_RE.match(token):
        uid = token
        member = await _fetch(uid)
        uname = getattr(member, "display_name", None) or getattr(member, "name", "") or uid
        return (uid, uname, member)

    # Plain username — search the guild member cache.
    if ctx.guild is not None:
        token_l = token.lower()
        for mem in getattr(ctx.guild, "members", []) or []:
            if (
                (getattr(mem, "name", "") or "").lower() == token_l
                or (getattr(mem, "display_name", "") or "").lower() == token_l
                or (getattr(mem, "global_name", "") or "").lower() == token_l
            ):
                return (str(mem.id), mem.display_name, mem)

    # Unresolvable — return the raw token as both id and name so the
    # caller can still emit a helpful error.
    return (token, token, None)


def _format_check_line(
    token: str,
    row: Optional[dict],
    today: date,
    events: Optional[list[dict]] = None,
) -> str:
    """Format a single mission_check line.

    If ``events`` is provided, append a short list of mission titles +
    fixer usernames pulled from the ``mission_event`` table so check
    output shows *what* missions a player has been on, not just *how
    many*.
    """
    events = events or []
    if not row and not events:
        return f"• **{token}** — no mission record."
    name = (row.get("username") if row else None) or token
    count = int((row.get("mission_count") if row else 0) or 0)
    dates = list((row.get("mission_dates") if row else None) or [])

    if dates:
        most_recent = max(dates)
        delta = (today - most_recent).days
        if delta < 0:
            when = f"future date {most_recent.isoformat()}"
        elif delta == 0:
            when = "today"
        elif delta == 1:
            when = "1 day ago"
        else:
            when = f"{delta} days ago"
        head = (
            f"• **{name}** — {count} mission(s); last on "
            f"**{most_recent.isoformat()}** ({when})."
        )
    elif count:
        head = f"• **{name}** — {count} mission(s), but no dates on file."
    else:
        head = f"• **{name}** — recent mission entries:"

    if not events:
        return head

    sub_lines: list[str] = []
    for e in events[:5]:
        mname = str(e.get("mission_name") or "Mission")
        if len(mname) > 60:
            mname = mname[:59] + "…"
        creator = str(e.get("creator_username") or "")
        if not creator:
            creator = f"<@{e.get('creator_id')}>" if e.get("creator_id") else "unknown"
        start_ts = e.get("start_ts")
        try:
            d_str = start_ts.date().isoformat()
        except Exception:
            d_str = ""
        sub_lines.append(
            f"   – *{mname}* (by **{creator}**" + (f", {d_str}" if d_str else "") + ")"
        )
    more = "" if len(events) <= 5 else f"\n   …and {len(events) - 5} more."
    return head + "\n" + "\n".join(sub_lines) + more


class MissionsCog(commands.Cog):
    """Mission tracking for the GM team."""

    PAYOUT_REASON = "NCRP Mission payout"

    def __init__(self, bot: commands.Bot, unbelievaboat=None):
        self.bot = bot
        self.unbelievaboat = unbelievaboat if unbelievaboat is not None else getattr(bot, "unbelievaboat", None)

    async def cog_load(self) -> None:
        self._payout_loop.start()

    async def cog_unload(self) -> None:
        self._payout_loop.cancel()

    @tasks.loop(minutes=5)
    async def _payout_loop(self) -> None:
        """Every 5 minutes, pay out any mission whose payout_ts has passed."""
        try:
            now_utc = datetime.now(timezone.utc)
            due = await mission_event_list_due(now_utc)
            for row in due:
                await self._process_mission_payout(row)
        except Exception:
            logger.error("mission _payout_loop iteration failed", exc_info=True)

    @_payout_loop.before_loop
    async def _before_payout_loop(self) -> None:
        await self.bot.wait_until_ready()

    async def _reconcile_mission_with_discord(self, row: dict) -> Optional[dict]:
        """Re-check the Discord scheduled event before paying.

        Returns:
          * ``row`` (possibly with synced fields) if the event still exists,
            its title still starts with ``Actors Needed:``, and its start
            time has already passed → caller should proceed with payout.
          * ``None`` if the mission should be skipped this cycle, either
            because:
              - the event was canceled/deleted (DB row marked canceled),
              - the title no longer matches an "Actors Needed:" event
                (DB row marked canceled),
              - the event was rescheduled into the future (DB row updated
                with the new start_ts/end_ts/payout_ts, loop will re-pick
                it on the new payout date),
              - Discord is temporarily unreachable (try again next loop).
        """
        mission_id = str(row.get("mission_id") or "")
        event_id = row.get("event_id")
        guild_id = row.get("guild_id")
        channel_id = row.get("channel_id")

        # Legacy / hand-crafted rows without a linked Discord event: skip
        # reconciliation and pay as before.
        if not event_id or not guild_id:
            return row

        guild = self.bot.get_guild(int(guild_id))
        if guild is None:
            try:
                guild = await self.bot.fetch_guild(int(guild_id))
            except Exception:
                logger.warning(
                    "Mission %s: guild %s unavailable; deferring payout",
                    mission_id, guild_id,
                )
                return None  # try again next loop iteration

        # Fetch the latest event state from Discord.
        event = None
        not_found = False
        try:
            event = await guild.fetch_scheduled_event(int(event_id))
        except discord.NotFound:
            not_found = True
        except Exception:
            logger.warning(
                "Mission %s: failed to fetch scheduled event %s; deferring",
                mission_id, event_id, exc_info=True,
            )
            return None  # transient — retry next loop

        if not_found or event is None:
            # Discord auto-purges completed scheduled events some time after
            # they end. If the mission's start_ts is in the past, treat a
            # NotFound as "the event ran and got cleaned up" — pay anyway.
            # Only cancel when the event was scheduled for the future and is
            # now missing (= fixer deleted it before it could run).
            start_ts = row.get("start_ts")
            now = datetime.now(timezone.utc)
            start_was_past = (
                isinstance(start_ts, datetime)
                and (start_ts if start_ts.tzinfo else start_ts.replace(tzinfo=timezone.utc)) <= now
            )
            if start_was_past:
                logger.info(
                    "Mission %s: Discord event %s not found but start_ts is "
                    "in the past — assuming completed-and-purged, paying anyway.",
                    mission_id, event_id,
                )
                return row
            await mission_event_cancel(mission_id)
            await self._notify_reconcile(
                row,
                title="🗑️ Mission Canceled (Event Deleted)",
                description=(
                    f"The Discord event for **{row.get('mission_name')}** "
                    "was deleted before it could run. No payout was issued."
                ),
                color=discord.Color.dark_grey(),
            )
            return None

        # Discord event status check (canceled by an organizer).
        ev_status = getattr(event, "status", None)
        try:
            is_canceled_status = ev_status == discord.EventStatus.canceled
        except Exception:
            is_canceled_status = str(ev_status).lower().endswith("canceled")
        if is_canceled_status:
            await mission_event_cancel(mission_id)
            await self._notify_reconcile(
                row,
                title="🗑️ Mission Canceled",
                description=(
                    f"The Discord event for **{row.get('mission_name')}** "
                    "was canceled. No payout was issued."
                ),
                color=discord.Color.dark_grey(),
            )
            return None

        # Only treat events tagged "Actors Needed:" as bot-managed missions.
        ev_name = getattr(event, "name", "") or ""
        if not ev_name.strip().lower().startswith(ACTORS_NEEDED_PREFIX.lower()):
            await mission_event_cancel(mission_id)
            await self._notify_reconcile(
                row,
                title="⚠️ Mission Skipped (Renamed)",
                description=(
                    f"The Discord event was renamed to `{ev_name[:120]}` and "
                    f"no longer begins with `{ACTORS_NEEDED_PREFIX}`. "
                    "Marking the mission canceled — no payout issued."
                ),
                color=discord.Color.orange(),
            )
            return None

        # Sync mutable fields. Title change → update mission_name.
        ev_start = getattr(event, "start_time", None) or getattr(event, "scheduled_start_time", None)
        ev_end = getattr(event, "end_time", None) or getattr(event, "scheduled_end_time", None)
        new_mission_name = _strip_actors_prefix(ev_name) or (row.get("mission_name") or "Mission")
        updates: dict = {}
        if new_mission_name and new_mission_name != row.get("mission_name"):
            updates["mission_name"] = new_mission_name
        if ev_start is not None and not _dt_close(ev_start, row.get("start_ts")):
            # Guardrail: if the event was moved more than 24 h into the past
            # (relative to NOW, not the original start), refuse to pay this
            # cycle. This blocks a misuse where someone rewinds an old event
            # to trigger a fresh payout.
            now = datetime.now(timezone.utc)
            ev_start_utc = ev_start if ev_start.tzinfo else ev_start.replace(tzinfo=timezone.utc)
            if ev_start_utc < now - timedelta(hours=24):
                await self._notify_reconcile(
                    row,
                    title="⚠️ Mission Skipped (Rewound >24h)",
                    description=(
                        f"The Discord event for **{row.get('mission_name')}** "
                        f"was moved to <t:{int(ev_start_utc.timestamp())}:F>, more "
                        "than 24 hours in the past. No payout will be issued and "
                        "the mission has been canceled. If this was intentional, "
                        "create a new mission for the correct date."
                    ),
                    color=discord.Color.orange(),
                )
                await mission_event_cancel(mission_id)
                return None
            updates["start_ts"] = ev_start
            if ev_end is not None:
                updates["end_ts"] = ev_end
                payout_basis = ev_end
            else:
                # Keep duration roughly consistent if Discord didn't return end.
                old_start = row.get("start_ts")
                old_end = row.get("end_ts")
                if isinstance(old_start, datetime) and isinstance(old_end, datetime):
                    inferred_end = ev_start + (old_end - old_start)
                    updates["end_ts"] = inferred_end
                    payout_basis = inferred_end
                else:
                    payout_basis = ev_start
            updates["payout_ts"] = compute_payout_ts(payout_basis)
        elif ev_end is not None and not _dt_close(ev_end, row.get("end_ts")):
            updates["end_ts"] = ev_end

        if updates:
            ok = await mission_event_update(mission_id, **updates)
            if not ok:
                logger.error(
                    "Mission %s: failed to apply Discord sync updates %r; deferring",
                    mission_id, list(updates.keys()),
                )
                return None
            # Reflect updates in the row dict for the payout step.
            row = dict(row)
            row.update(updates)
            # If the *new* payout time is in the future, defer this cycle —
            # the loop will re-pick the row when payout_ts arrives. Key off
            # payout_ts (not start_ts) because someone could move start into
            # the past but the recomputed midnight-ET payout still be future.
            new_payout_ts = updates.get("payout_ts", row.get("payout_ts"))
            new_start_ts = updates.get("start_ts", row.get("start_ts"))
            now = datetime.now(timezone.utc)
            if isinstance(new_payout_ts, datetime):
                pts = new_payout_ts if new_payout_ts.tzinfo else new_payout_ts.replace(tzinfo=timezone.utc)
                if pts > now:
                    display_when = new_start_ts if isinstance(new_start_ts, datetime) else new_payout_ts
                    if display_when.tzinfo is None:
                        display_when = display_when.replace(tzinfo=timezone.utc)
                    await self._notify_reconcile(
                        row,
                        title="📅 Mission Rescheduled",
                        description=(
                            f"**{new_mission_name}** was moved on Discord to "
                            f"<t:{int(display_when.timestamp())}:F>. Payout deferred "
                            f"to <t:{int(pts.timestamp())}:F>."
                        ),
                        color=discord.Color.blurple(),
                    )
                    return None

        return row

    async def _notify_reconcile(
        self,
        row: dict,
        *,
        title: str,
        description: str,
        color: discord.Color,
    ) -> None:
        """Best-effort post a reconciliation notice to the mission's channel."""
        guild_id = row.get("guild_id")
        channel_id = row.get("channel_id")
        if not (guild_id and channel_id):
            return
        try:
            guild = self.bot.get_guild(int(guild_id))
            ch = guild.get_channel(int(channel_id)) if guild else None
            if ch is None:
                return
            embed = discord.Embed(title=title, description=description, color=color)
            creator_id = row.get("creator_id")
            if creator_id:
                embed.add_field(name="Fixer", value=f"<@{creator_id}>", inline=True)
            await ch.send(embed=embed)
        except Exception:
            logger.warning(
                "Failed to post reconcile notice for mission %s",
                row.get("mission_id"), exc_info=True,
            )

    async def _process_mission_payout(self, row: dict) -> None:
        # Re-fetch the row so we pay with the latest pay_per_player /
        # attendee_ids / canceled state, not whatever was snapshotted when
        # the loop started.
        mid = str(row.get("mission_id") or "")
        if mid:
            fresh = await mission_event_get(mid)
            if fresh is None:
                logger.info("Mission %s no longer exists; skipping payout", mid)
                return
            if fresh.get("paid"):
                # Already paid in a previous loop iteration.
                return
            if fresh.get("canceled"):
                logger.info("Mission %s is canceled; skipping payout", mid)
                return
            row = fresh

        # Reconcile against Discord — skip / reschedule / cancel as needed
        # before touching wallets.
        reconciled = await self._reconcile_mission_with_discord(row)
        if reconciled is None:
            return
        row = reconciled

        # Atomically claim the row before any UB call. This single CAS
        # closes the double-payout race: if two loop ticks (or a concurrent
        # cancel) interleave, exactly one wins the claim and the other
        # no-ops.
        if mid:
            claimed = await mission_event_claim_for_payout(mid)
            if claimed is None:
                logger.info(
                    "Mission %s: claim_for_payout returned None — already paid "
                    "or canceled in the meantime; skipping.",
                    mid,
                )
                return
            row = claimed

        mission_id = row.get("mission_id")
        mission_name = row.get("mission_name") or "Mission"
        pay = int(row.get("pay_per_player") or 0)
        attendees = list(row.get("attendee_ids") or [])
        guild_id = row.get("guild_id")
        channel_id = row.get("channel_id")
        start_ts = row.get("start_ts")
        mission_date = start_ts.date() if isinstance(start_ts, datetime) else datetime.now(timezone.utc).date()

        paid_lines: list[str] = []
        failed_lines: list[str] = []
        paid_uids: list[str] = []

        for uid in attendees:
            display = uid
            try:
                user = await self.bot.fetch_user(int(uid))
                display = getattr(user, "display_name", None) or getattr(user, "name", uid)
            except Exception:
                user = None

            ub_ok = True
            if pay > 0 and self.unbelievaboat is not None:
                try:
                    ub_ok = await self.unbelievaboat.update_balance(
                        int(uid),
                        {"cash": 0, "bank": pay},
                        reason=f"{self.PAYOUT_REASON}: {mission_name}",
                    )
                except Exception:
                    logger.error("UB payout failed for %s on mission %s", uid, mission_id, exc_info=True)
                    ub_ok = False

            # mission_log was already recorded at creation time, so the
            # payout path only credits UB and reports results.
            if ub_ok:
                paid_lines.append(f"• **{display}** — +¥{pay:,} bank")
                paid_uids.append(str(uid))
            else:
                failed_lines.append(f"• **{display}** — UB payout failed")

        # Row was already claimed (paid=TRUE) at the start of this function
        # via mission_event_claim_for_payout. Legacy rows that skipped the
        # claim (no mission_id) still need the old mark_paid path.
        if not mid:
            await mission_event_mark_paid(str(mission_id))

        # Post a summary back to the channel where the mission was created.
        if guild_id and channel_id:
            try:
                guild = self.bot.get_guild(int(guild_id))
                ch = guild.get_channel(int(channel_id)) if guild else None
                if ch is not None:
                    embed = discord.Embed(
                        title=f"💰 Mission Payout — {mission_name}",
                        color=discord.Color.gold(),
                    )
                    if paid_lines:
                        embed.add_field(
                            name=f"Paid ({len(paid_lines)})",
                            value="\n".join(paid_lines)[:1024],
                            inline=False,
                        )
                    if failed_lines:
                        embed.add_field(
                            name=f"Failed ({len(failed_lines)})",
                            value="\n".join(failed_lines)[:1024],
                            inline=False,
                        )
                    if not paid_lines and not failed_lines:
                        embed.description = "No attendees on this mission — nothing to pay."
                    await ch.send(embed=embed)
            except Exception:
                logger.error("Failed to post mission payout summary for %s", mission_id, exc_info=True)

        creator_id = row.get("creator_id")
        creator_username = row.get("creator_username") or (str(creator_id) if creator_id else None)

        # Post a single per-mission summary to #player-gig-record so players
        # have a public ledger of who got paid what for which gig.
        if paid_uids:
            try:
                import config as _cfg
                gig_ch_id = getattr(_cfg, "PLAYER_GIG_RECORD_CHANNEL_ID", 0)
                if gig_ch_id:
                    gig_ch = self.bot.get_channel(int(gig_ch_id))
                    if gig_ch is None:
                        try:
                            gig_ch = await self.bot.fetch_channel(int(gig_ch_id))
                        except Exception:
                            gig_ch = None
                    if gig_ch is not None:
                        date_str = (
                            mission_date.isoformat()
                            if mission_date else "_(unknown)_"
                        )
                        fixer_part = (
                            f"<@{creator_id}>" if creator_id
                            else (str(creator_username) if creator_username else "_(unknown)_")
                        )
                        player_lines = "\n".join(
                            f"• <@{uid}> — **¥{pay:,}** (bank)" for uid in paid_uids
                        )
                        msg = (
                            f"💰 **Gig Payout** — _{mission_name}_\n"
                            f"📅 **Date:** {date_str}\n"
                            f"🎬 **Fixer:** {fixer_part}\n"
                            f"{player_lines}"
                        )
                        await gig_ch.send(
                            msg,
                            allowed_mentions=discord.AllowedMentions(
                                users=True, roles=False, everyone=False
                            ),
                        )
            except Exception:
                logger.error(
                    "Failed to post gig-record summary for mission %s",
                    mission_id, exc_info=True,
                )

        await _post_mission_audit(
            self.bot,
            action="Mission Auto-Payout",
            actor_id=str(creator_id) if creator_id else None,
            actor_name=str(creator_username) if creator_username else None,
            mission_id=str(mission_id) if mission_id else None,
            mission_name=mission_name,
            fields=[
                ("Pay per attendee", f"¥{pay:,} → bank"),
                ("Mission date", mission_date.isoformat() if mission_date else "_(unknown)_"),
                (
                    f"Paid ({len(paid_lines)})",
                    "\n".join(paid_lines) if paid_lines else "_(none)_",
                ),
                (
                    f"Failed ({len(failed_lines)})",
                    "\n".join(failed_lines) if failed_lines else "_(none)_",
                ),
            ],
            color=discord.Color.gold(),
        )

    @commands.command(name="mission_check", aliases=["mission_viability"])
    @is_fixer()
    async def mission_check(self, ctx: commands.Context, *targets: str):
        """Show mission history for each given user (mention, ID, or username)."""
        if not targets:
            await ctx.reply(
                "Usage: `!mission_check @user [@user2 …]` — accepts mentions, "
                "Discord IDs, or usernames.",
                mention_author=False,
            )
            return
        today = datetime.now(timezone.utc).date()
        lines: list[str] = []
        for tok in targets:
            uid, uname, _user_obj = await _resolve_target(ctx, tok)
            if not uid:
                continue
            row = await mission_log_get(uid)
            events = await mission_event_get_for_user(uid, limit=25)
            display_token = uname or tok
            lines.append(_format_check_line(display_token, row, today, events))
        if not lines:
            await ctx.reply("No valid targets given.", mention_author=False)
            return
        await ctx.reply("\n".join(lines), mention_author=False)

    @commands.command(name="mission_record")
    @is_fixer()
    async def mission_record(self, ctx: commands.Context, *args: str):
        """Record a mission for each user. Optional ``date=YYYY-MM-DD``."""
        targets: list[str] = []
        mission_date = datetime.now(timezone.utc).date()
        explicit_date = False
        for arg in args:
            if arg.lower().startswith("date="):
                parsed = _parse_date_token(arg.split("=", 1)[1])
                if parsed is None:
                    await ctx.reply(
                        f"Couldn't parse date `{arg}`. Use `date=YYYY-MM-DD` "
                        "or `date=M/D/YYYY`.",
                        mention_author=False,
                    )
                    return
                mission_date = parsed
                explicit_date = True
            else:
                targets.append(arg)
        if not targets:
            await ctx.reply(
                "Usage: `!mission_record @user [@user2 …] [date=YYYY-MM-DD]`",
                mention_author=False,
            )
            return
        recorded: list[str] = []
        failed: list[str] = []
        for tok in targets:
            uid, uname, _user_obj = await _resolve_target(ctx, tok)
            if not uid:
                continue
            result = await mission_log_record(uid, uname or "", mission_date)
            if result is None:
                failed.append(tok)
                continue
            recorded.append(
                f"• **{result.get('username') or uname or uid}** → "
                f"total {int(result.get('mission_count') or 0)} mission(s)"
            )
        date_label = mission_date.isoformat() + (" (today)" if not explicit_date else "")
        body = f"Recorded mission on **{date_label}**:\n" + "\n".join(recorded)
        if failed:
            body += "\nFailed: " + ", ".join(f"`{t}`" for t in failed)
        await ctx.reply(body, mention_author=False)
        recorded_ids: list[str] = []
        for tok in targets:
            uid, _uname, _user = await _resolve_target(ctx, tok)
            if uid:
                recorded_ids.append(str(uid))
        await _post_mission_audit(
            self.bot,
            action="Mission Recorded (command)",
            actor=ctx.author,
            fields=[
                ("Date", mission_date.isoformat()),
                (
                    f"Players ({len(recorded_ids)})",
                    " ".join(f"<@{uid}>" for uid in recorded_ids) if recorded_ids else "_(none)_",
                ),
                (
                    f"Failed ({len(failed)})",
                    ", ".join(f"`{t}`" for t in failed) if failed else "_(none)_",
                ),
            ],
            color=discord.Color.green(),
        )

async def setup(bot: commands.Bot):
    await bot.add_cog(MissionsCog(bot))
