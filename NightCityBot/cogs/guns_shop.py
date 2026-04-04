"""Gun shop cog — two-tier wholesale-to-store supply chain.

The legacy prefix commands (!guns_wh_setshop, !guns_wh_shops, !guns_wh_list,
!guns_store_inv, !guns_wh_buy, !guns_wh_sell, !guns_wh_restock,
!guns_wh_clear_inventory, !guns_wh_restock_settings, !guns_wh_setsheet,
!guns_wh_recheck, !guns_wh_gunlist, !guns_wh_paths, !guns_wh_add,
!guns_store_add, !guns_wh_remove, !guns_store_remove, !guns_wh_approve,
!guns_wh_unapprove, !guns_wh_approved, !guns_wh_tx, !guns_wh_retry_payout)
have been removed.  All gun-shop actions are now handled through the
Gunstore Hub (!gunstore) and Fixer Hub (!fixer).

This cog is still loaded so that hub code can access the helper methods
(state management, restock logic, sheet parsing, balance helpers) via
``bot.cogs.get("GunsShopCog")``.
"""
import asyncio
import json
import logging
import random
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional
from urllib.parse import urlparse, parse_qs

import aiohttp
import discord
from discord.ext import commands, tasks
from openpyxl import load_workbook

import config
from NightCityBot.utils import helpers
from NightCityBot.utils.db import (
    wh_lots_get_all, wh_lots_replace_all,
    wh_stores_get_all, wh_stores_replace_all,
    wh_shops_get_all, wh_shops_replace_all,
    wh_settings_get, wh_settings_save,
    wh_tx_append, wh_tx_get_all,
    gun_catalog_upsert_many, gun_catalog_sync_qty_from_lots, gun_catalog_adjust_qty,
    gun_catalog_get_all,
    pi_add_item,
    ResourceLockManager,
)

logger = logging.getLogger(__name__)


class GunsShopCog(commands.Cog):
    """Two-tier gun supply chain: corp wholesaler -> stores -> players.

    Uses read-only spreadsheet parsing and immutable receipt/audit logs.
    Staff updates Character Gun Tracking manually from receipts.
    """

    LEVEL_SETTINGS = {
        "L": {"weight": 70, "qty_min": 3, "qty_max": 10},
        "M": {"weight": 25, "qty_min": 1, "qty_max": 5},
        "H": {"weight": 5, "qty_min": 1, "qty_max": 2},
    }
    DEFAULT_RESTOCK_SETTINGS = {
        "total_lots": 20,
        "lots_L": 14,
        "lots_M": 5,
        "lots_H": 1,
        "qty_min_L": 3,
        "qty_max_L": 10,
        "qty_min_M": 1,
        "qty_max_M": 5,
        "qty_min_H": 1,
        "qty_max_H": 2,
    }
    WEAPON_TYPES = (
        "pistol",
        "revolver",
        "submachine_gun",
        "shotgun",
        "assault_rifle",
        "light_machine_gun",
        "heavy_machine_gun",
        "precision_rifle",
        "sniper_rifle",
    )
    VALID_RESTRICTIONS = ("basic", "controlled", "restricted")
    WEAPON_TYPE_PATTERNS = {
        "pistol": ("pistol",),
        "revolver": ("revolver",),
        "submachine_gun": ("submachine gun", "submachine-gun", "smg"),
        "shotgun": ("shotgun",),
        "assault_rifle": ("assault rifle", "assault-rifle", "ar "),
        "light_machine_gun": ("light machine gun", "light-machine-gun", "lmg"),
        "heavy_machine_gun": ("heavy machine gun", "heavy-machine-gun", "hmg"),
        "precision_rifle": ("precision rifle", "precision-rifle", "dmr"),
        "sniper_rifle": ("sniper rifle", "sniper-rifle", "sniper"),
    }

    def __init__(self, bot: commands.Bot) -> None:
        self.bot = bot
        self.unbelievaboat = bot.unbelievaboat

        base_dir = Path(getattr(config, "BASE_DIR", Path(__file__).resolve().parents[1]))
        self.base_dir = base_dir
        default_data_dir = base_dir / "data" / "wholesaler"
        configured_data_dir = getattr(config, "WHOLESALER_DATA_DIR", None)
        self.data_dir = self._resolve_base_path(configured_data_dir, default_data_dir)
        self.state_file = self._resolve_data_path(
            getattr(config, "WHOLESALER_STATE_FILE", None),
            "state.json",
        )

        self.data_dir.mkdir(parents=True, exist_ok=True)

        self.sheet_cache_path = self._resolve_data_path(
            getattr(config, "WHOLESALER_SHEET_CACHE_FILE", None),
            "master_sheet_latest.xlsx",
        )
        self.store_state_file = self._resolve_data_path(
            getattr(config, "WHOLESALER_STORE_STATE_FILE", None),
            "stores.json",
        )
        self.wholesale_inventory_file = self._resolve_data_path(
            getattr(config, "WHOLESALER_WHOLESALE_FILE", None),
            "inventory/wholesale.json",
        )
        self.store_inventory_dir = self._resolve_data_path(
            getattr(config, "WHOLESALER_STORE_INVENTORY_DIR", None),
            "inventory/stores",
        )
        self.tx_file = self._resolve_data_path(
            getattr(config, "WHOLESALER_TX_FILE", None),
            "transactions.json",
        )

        self._migrate_legacy_files(default_data_dir)
        self.wholesale_inventory_file.parent.mkdir(parents=True, exist_ok=True)
        self.store_inventory_dir.mkdir(parents=True, exist_ok=True)
        self._locks = ResourceLockManager()
        self.lock = self._locks.pin("state")
        self._startup_audit_sent = False
        self.weekly_sunday_restock.start()

    def _resolve_data_path(self, configured_path: Any, default_relative: str) -> Path:
        if configured_path is None:
            return self.data_dir / default_relative

        configured = Path(str(configured_path)).expanduser()
        if configured.is_absolute():
            return configured
        return self.data_dir / configured

    def _resolve_base_path(self, configured_path: Any, default_path: Path) -> Path:
        """Resolve base-level paths, anchoring relative values to config.BASE_DIR."""
        if configured_path is None:
            return default_path

        configured = Path(str(configured_path)).expanduser()
        if configured.is_absolute():
            return configured
        return self.base_dir / configured

    def _migrate_legacy_files(self, legacy_dir: Path) -> None:
        """Copy old in-repo wholesaler files into configured persistent paths once."""
        if self.data_dir == legacy_dir:
            return

        store_state_file = getattr(self, "store_state_file", self.state_file)

        migrations = (
            (legacy_dir / "state.json", self.state_file),
            (legacy_dir / "state.json", store_state_file),
            (legacy_dir / "transactions.json", self.tx_file),
            (legacy_dir / "master_sheet_latest.xlsx", self.sheet_cache_path),
        )
        for src, dst in migrations:
            if dst.exists() or not src.exists():
                continue
            dst.parent.mkdir(parents=True, exist_ok=True)
            try:
                dst.write_bytes(src.read_bytes())
                logger.info("Migrated wholesaler data file %s -> %s", src, dst)
            except Exception:
                logger.exception("Failed to migrate wholesaler data file %s -> %s", src, dst)

        legacy_wholesale = legacy_dir / "inventory" / "wholesale.json"
        if legacy_wholesale.exists() and not self.wholesale_inventory_file.exists():
            self.wholesale_inventory_file.parent.mkdir(parents=True, exist_ok=True)
            try:
                self.wholesale_inventory_file.write_bytes(legacy_wholesale.read_bytes())
                logger.info("Migrated wholesaler inventory file %s -> %s", legacy_wholesale, self.wholesale_inventory_file)
            except Exception:
                logger.exception("Failed to migrate wholesaler inventory file %s -> %s", legacy_wholesale, self.wholesale_inventory_file)

        legacy_store_dir = legacy_dir / "inventory" / "stores"
        if legacy_store_dir.exists():
            self.store_inventory_dir.mkdir(parents=True, exist_ok=True)
            for src in sorted(legacy_store_dir.glob("*.json")):
                dst = self.store_inventory_dir / src.name
                if dst.exists():
                    continue
                try:
                    dst.write_bytes(src.read_bytes())
                    logger.info("Migrated store inventory file %s -> %s", src, dst)
                except Exception:
                    logger.exception("Failed to migrate store inventory file %s -> %s", src, dst)


    def _store_inventory_file(self, store_id: str) -> Path:
        safe_store_id = re.sub(r"[^a-zA-Z0-9_-]+", "-", str(store_id)).strip("-")
        store_inventory_dir = Path(getattr(self, "store_inventory_dir", Path(getattr(self, "state_file", Path("state.json"))).parent / "inventory" / "stores"))
        return store_inventory_dir / f"{safe_store_id or 'store'}.json"

    def _list_store_inventory_files(self) -> list[Path]:
        store_inventory_dir = Path(getattr(self, "store_inventory_dir", Path(getattr(self, "state_file", Path("state.json"))).parent / "inventory" / "stores"))
        if not store_inventory_dir.exists():
            return []
        return sorted(store_inventory_dir.glob("*.json"))

    def cog_unload(self):
        self.weekly_sunday_restock.cancel()

    @tasks.loop(hours=1)
    async def weekly_sunday_restock(self) -> None:
        """Automatically refresh wholesaler stock every Sunday (UTC)."""
        now = datetime.now(timezone.utc)
        if now.weekday() != 6:
            return
        await self._auto_restock_if_due(now, trigger="SCHEDULED")

    @weekly_sunday_restock.before_loop
    async def before_weekly_sunday_restock(self) -> None:
        await self.bot.wait_until_ready()

    @staticmethod
    def _now_iso() -> str:
        return datetime.now(timezone.utc).isoformat()

    @staticmethod
    def _to_int(value: Any) -> Optional[int]:
        if value is None:
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        stripped = str(value).replace(",", "").replace("$", "").strip()
        if not stripped:
            return None
        try:
            return int(float(stripped))
        except ValueError:
            return None

    @staticmethod
    def _derive_level(effectiveness_raw: str) -> str:
        text = (effectiveness_raw or "").upper()
        if "(M-H)" in text or "(H)" in text:
            return "H"
        if "(M)" in text:
            return "M"
        if "(L)" in text:
            return "L"
        return "L"

    @staticmethod
    def _derive_category(effectiveness_raw: str) -> Optional[str]:
        text = (effectiveness_raw or "").lower()
        for category in ("power", "tech", "smart"):
            if category in text:
                return category.title()
        return None

    @classmethod
    def _derive_weapon_type(cls, gun_name: str, effectiveness_raw: str) -> Optional[str]:
        haystack = f"{gun_name or ''} {effectiveness_raw or ''}".lower()
        for weapon_type, patterns in cls.WEAPON_TYPE_PATTERNS.items():
            if any(pattern in haystack for pattern in patterns):
                return weapon_type
        return None

    @staticmethod
    def _normalize_shop_name(name: str) -> str:
        cleaned = re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-")
        return cleaned

    @staticmethod
    def _inventory_totals(state: dict[str, Any]) -> tuple[int, int, int, int]:
        wholesale_lots = state.get("wholesale_lots", [])
        wholesale_lot_count = len(wholesale_lots) if isinstance(wholesale_lots, list) else 0
        wholesale_unit_count = (
            sum(max(int(lot.get("qty_available", 0)), 0) for lot in wholesale_lots)
            if isinstance(wholesale_lots, list)
            else 0
        )

        stores = state.get("stores", {})
        store_count = len(stores) if isinstance(stores, dict) else 0
        store_unit_count = 0
        if isinstance(stores, dict):
            for payload in stores.values():
                if not isinstance(payload, dict):
                    continue
                for lot in payload.get("lots", []):
                    if isinstance(lot, dict):
                        store_unit_count += max(int(lot.get("qty_remaining", 0)), 0)

        return wholesale_lot_count, wholesale_unit_count, store_count, store_unit_count

    async def emit_inventory_snapshot_audit(
        self,
        reason: str,
        *,
        actor: Optional[discord.abc.User] = None,
    ) -> None:
        """Emit an audit line that captures persisted inventory state."""
        try:
            state = await self._load_state()
            lots, wholesale_units, stores, store_units = self._inventory_totals(state)
            wholesale_exists = self.wholesale_inventory_file.exists()
            wholesale_bytes = self.wholesale_inventory_file.stat().st_size if wholesale_exists else 0
            state_exists = self.state_file.exists()
            state_bytes = self.state_file.stat().st_size if state_exists else 0
            actor_text = f" actor=<@{actor.id}>" if actor else ""
            await self._audit_send(
                "[WHOLESALE_SNAPSHOT]"
                f" reason={reason}{actor_text}"
                f" lots={lots} wholesale_units={wholesale_units}"
                f" stores={stores} store_units={store_units}"
                f" wholesale_file={self.wholesale_inventory_file}"
                f" wholesale_file_exists={wholesale_exists} wholesale_file_bytes={wholesale_bytes}"
                f" state_file={self.state_file}"
                f" state_file_exists={state_exists} state_file_bytes={state_bytes}"
            )
        except Exception:
            logger.exception("Failed to emit wholesaler inventory snapshot for reason=%s", reason)

    @commands.Cog.listener()
    async def on_ready(self) -> None:
        if self._startup_audit_sent:
            return
        self._startup_audit_sent = True
        logger.info(
            "Wholesaler data paths: data_dir=%s state=%s wholesale=%s stores=%s store_inv_dir=%s tx=%s",
            self.data_dir,
            self.state_file,
            self.wholesale_inventory_file,
            self.store_state_file,
            self.store_inventory_dir,
            self.tx_file,
        )
        await self._ensure_inventory_files_exist()
        logger.info(
            "Wholesaler files verified: state=%s wholesale=%s stores=%s tx=%s",
            self.state_file.exists(),
            self.wholesale_inventory_file.exists(),
            self.store_state_file.exists(),
            self.tx_file.exists(),
        )
        await wh_tx_get_all()
        await self.emit_inventory_snapshot_audit("BOT_READY")

        # Populate gun_catalog on startup if the table is empty.
        # This handles the case where the bot restarts after the table was first
        # added but before any restock has been run.
        try:
            existing = await gun_catalog_get_all()
            if not existing:
                logger.info("gun_catalog is empty — attempting to populate from sheet on startup")
                guns = await self._load_master_guns()
                if guns:
                    state = await self._load_state()
                    lots = state.get("wholesale_lots", [])
                    await gun_catalog_sync_qty_from_lots(lots)
                    logger.info(
                        "gun_catalog populated on startup: %d guns, %d lots synced",
                        len(guns), len(lots),
                    )
                else:
                    logger.info("gun_catalog startup populate: no guns found (sheet not configured?)")
            else:
                logger.info("gun_catalog already populated (%d entries) — skipping startup reload", len(existing))
        except Exception:
            logger.warning("gun_catalog startup populate failed (non-fatal)", exc_info=True)

    async def _ensure_inventory_files_exist(self) -> None:
        """Create baseline wholesaler/store persistence files during startup."""
        async with self.lock:
            data_dir = Path(getattr(self, "data_dir", self.state_file.parent))
            # Create expected files/directories up front so deployments can
            # validate storage locations immediately after startup.
            data_dir.mkdir(parents=True, exist_ok=True)
            self.wholesale_inventory_file.parent.mkdir(parents=True, exist_ok=True)
            self.store_inventory_dir.mkdir(parents=True, exist_ok=True)

            if not self.state_file.exists():
                await helpers.save_json_file(
                    self.state_file,
                    {
                        "wholesale_lots": [],
                        "transactions": 0,
                        "pending_payouts": [],
                        "settings": {},
                    },
                )
            if not self.store_state_file.exists():
                await helpers.save_json_file(
                    self.store_state_file,
                    {
                        "shop_registry": {},
                        "stores": {},
                    },
                )
            if not self.wholesale_inventory_file.exists():
                await helpers.save_json_file(
                    self.wholesale_inventory_file,
                    {
                        "wholesale_lots": [],
                    },
                )
            if not self.tx_file.exists():
                await helpers.save_json_file(self.tx_file, [])

            state = await self._load_state()
            saved = await self._save_state(state)
            if not saved:
                logger.warning(
                    "Wholesaler startup bootstrap could not persist one or more inventory files "
                    "(state=%s, wholesale=%s, store_index=%s, store_dir=%s)",
                    self.state_file,
                    self.wholesale_inventory_file,
                    self.store_state_file,
                    self.store_inventory_dir,
                )
            base_dir = Path(getattr(self, "base_dir", data_dir.parent))
            if data_dir != base_dir / "data" / "wholesaler":
                logger.info(
                    "Wholesaler data directory override active data_dir=%s base_dir=%s",
                    data_dir,
                    base_dir,
                )


    @staticmethod
    def _clean_sheet_url_input(value: str) -> str:
        """Normalize user-provided URL text from Discord messages."""
        raw = (value or "").strip()
        if raw.startswith("<") and raw.endswith(">"):
            raw = raw[1:-1].strip()
        return raw

    @staticmethod
    def _normalize_sheet_source_url(value: str) -> str:
        """Normalize a Google Sheets URL to an XLSX export URL when possible."""
        raw = GunsShopCog._clean_sheet_url_input(value)
        if not raw:
            return raw
        parsed = urlparse(raw)

        if "docs.google.com" not in parsed.netloc:
            return raw

        # Already an export endpoint
        if "/export" in parsed.path:
            return raw

        match = re.search(r"/spreadsheets/(?:u/\d+/)?d/([a-zA-Z0-9-_]+)", parsed.path)
        if not match:
            return raw

        sheet_id = match.group(1)
        gid = parse_qs(parsed.query).get("gid", [None])[0]
        if not gid and parsed.fragment:
            frag_qs = parse_qs(parsed.fragment)
            gid = frag_qs.get("gid", [None])[0]
            if not gid and parsed.fragment.startswith("gid="):
                gid = parsed.fragment.split("=", 1)[1].split("&", 1)[0]

        export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        if gid:
            export_url += f"&gid={gid}"
        return export_url

    @staticmethod
    def _extract_sheet_gid(value: str) -> Optional[str]:
        """Extract a Google Sheets gid from a URL query or fragment."""
        raw = GunsShopCog._clean_sheet_url_input(value)
        if not raw:
            return None
        parsed = urlparse(raw)
        gid = parse_qs(parsed.query).get("gid", [None])[0]
        if gid:
            return gid
        if parsed.fragment:
            frag_qs = parse_qs(parsed.fragment)
            gid = frag_qs.get("gid", [None])[0]
            if gid:
                return gid
            if parsed.fragment.startswith("gid="):
                return parsed.fragment.split("=", 1)[1].split("&", 1)[0]
        return None

    @staticmethod
    def parse_master_sheet(
        xlsx_path: str | Path,
        sheet_name: str,
        sheet_gid: Optional[str] = None,
    ) -> list[dict[str, Any]]:
        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
        selected_sheet = sheet_name if sheet_name in wb.sheetnames else None
        if not selected_sheet and sheet_gid:
            for candidate in wb.worksheets:
                tab_id = getattr(candidate.sheet_properties, "tabId", None)
                if tab_id is not None and str(tab_id) == str(sheet_gid):
                    selected_sheet = candidate.title
                    break

        # Some exports only include one visible tab. Use it as a safe fallback.
        if not selected_sheet and len(wb.sheetnames) == 1:
            selected_sheet = wb.sheetnames[0]

        if not selected_sheet:
            available = ", ".join(wb.sheetnames)
            wb.close()
            raise ValueError(
                f"Sheet '{sheet_name}' not found in workbook"
                f" (available: {available})"
            )

        ws = wb[selected_sheet]
        row_iter = ws.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if not header_row:
            wb.close()
            return []

        header = [str(c).strip() if c is not None else "" for c in header_row]
        header_lookup = {name.lower(): idx for idx, name in enumerate(header)}

        def idx_for(options: list[str], fallback: int) -> int:
            for opt in options:
                found = header_lookup.get(opt.lower())
                if found is not None:
                    return found
            return fallback

        name_idx = idx_for(["Gun Name", "Name", "Weapon"], 0)
        eff_idx = idx_for(["Type/Armor Effectiveness", "Type", "Effectiveness"], 1)
        mag_idx = idx_for(["Mag Size", "Mag"], 2)
        price_idx = idx_for(["Price New", "Price", "Price (New)"], 3)
        cyberware_idx = idx_for(["Cyberware Needed", "Cyberware"], 4)
        restriction_idx = idx_for(["Restriction", "Restrictions"], None)
        status_idx = idx_for(["Status", "Market Status"], None)
        _raw_type_idx = idx_for(["Type"], None)
        type_idx = _raw_type_idx if (_raw_type_idx is not None and _raw_type_idx != eff_idx) else None
        power_level_idx = idx_for(["Power Level", "PowerLevel", "Power_Level"], None)

        section_header_map = {
            "pistols": "pistol",
            "revolvers": "revolver",
            "submachine guns": "submachine_gun",
            "shotguns": "shotgun",
            "assault rifles": "assault_rifle",
            "light machine guns": "light_machine_gun",
            "heavy machine guns": "heavy_machine_gun",
            "precision rifles": "precision_rifle",
            "sniper rifles": "sniper_rifle",
        }
        current_section_type = "pistol"

        parsed: list[dict[str, Any]] = []
        for row in row_iter:
            if not row:
                continue

            gun_name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] is not None else ""
            effectiveness_raw = (
                str(row[eff_idx]).strip() if eff_idx < len(row) and row[eff_idx] is not None else ""
            )
            price_new = GunsShopCog._to_int(row[price_idx] if price_idx < len(row) else None)

            if not gun_name:
                continue
            if effectiveness_raw.lower() == "type":
                matched_section = section_header_map.get(gun_name.lower().strip())
                if matched_section:
                    current_section_type = matched_section
                    logger.debug("Sheet section header: '%s' → weapon_type=%s", gun_name, current_section_type)
                continue
            if price_new is None or price_new <= 0:
                continue

            weapon_type = GunsShopCog._derive_weapon_type(gun_name, effectiveness_raw) or current_section_type

            mag_raw = row[mag_idx] if mag_idx < len(row) else None
            mag_size = GunsShopCog._to_int(mag_raw)
            if mag_size is None and mag_raw is not None:
                mag_size = str(mag_raw)

            cyber_raw = row[cyberware_idx] if cyberware_idx < len(row) else ""
            cyberware_needed = GunsShopCog._to_int(cyber_raw)
            if cyberware_needed is None:
                cyberware_needed = "" if cyber_raw is None else str(cyber_raw)

            restriction_raw = ""
            if restriction_idx is not None and restriction_idx < len(row) and row[restriction_idx] is not None:
                restriction_raw = str(row[restriction_idx]).strip().lower()
            restriction = restriction_raw if restriction_raw in ("basic", "controlled", "restricted") else "basic"

            status = "live"
            if status_idx is not None and status_idx < len(row) and row[status_idx] is not None:
                status = str(row[status_idx]).strip().lower()

            direct_type = ""
            if type_idx is not None and type_idx < len(row) and row[type_idx] is not None:
                direct_type = str(row[type_idx]).strip().lower()
            gun_category = direct_type.title() if direct_type in ("power", "smart", "tech") else GunsShopCog._derive_category(effectiveness_raw)

            direct_pl = ""
            if power_level_idx is not None and power_level_idx < len(row) and row[power_level_idx] is not None:
                direct_pl = str(row[power_level_idx]).strip().lower()
            pl_map = {"low": "L", "medium": "M", "high": "H", "l": "L", "m": "M", "h": "H"}
            gun_level = pl_map.get(direct_pl, "") or GunsShopCog._derive_level(effectiveness_raw)

            parsed.append(
                {
                    "gun_name": gun_name,
                    "effectiveness_raw": effectiveness_raw,
                    "mag_size": mag_size,
                    "price_new": price_new,
                    "cyberware_needed": cyberware_needed,
                    "gun_level": gun_level,
                    "gun_category": gun_category,
                    "weapon_type": weapon_type,
                    "restriction": restriction,
                    "status": status,
                }
            )

        wb.close()
        return parsed

    async def _resolve_sheet_source(self) -> tuple[str, Optional[str]]:
        """Resolve active master-sheet URL and optional gid override."""
        state = await self._load_state()
        configured_url = str(state.get("settings", {}).get("master_sheet_url", "")).strip()
        config_url = str(getattr(config, "WHOLESALER_GOOGLE_SHEET_XLSX_URL", "") or "").strip()
        raw_url = configured_url or config_url

        local_fallback = str(getattr(config, "WHOLESALER_XLSX_PATH", "") or "").strip()
        if not raw_url and local_fallback.startswith(("http://", "https://")):
            raw_url = local_fallback

        return self._normalize_sheet_source_url(raw_url), self._extract_sheet_gid(raw_url)

    async def _resolve_sheet_path(self) -> Path:
        """Return local xlsx path, downloading from Google Sheets if configured."""
        state = await self._load_state()
        configured_url = str(state.get("settings", {}).get("master_sheet_url", "")).strip()
        sheet_url, _ = await self._resolve_sheet_source()
        if not sheet_url:
            local_path = Path(config.WHOLESALER_XLSX_PATH)
            if not local_path.exists():
                raise FileNotFoundError(
                    f"Local wholesaler sheet not found at '{local_path}'. "
                    "Set a Google Sheet source via the `!admin` hub, or update WHOLESALER_XLSX_PATH."
                )
            return local_path

        # Self-heal older stored share URLs by persisting normalized export URL.
        if configured_url and configured_url != sheet_url:
            async with self.lock:
                latest = await self._load_state()
                latest.setdefault("settings", {})["master_sheet_url"] = sheet_url
                await self._save_state(latest)

        timeout = aiohttp.ClientTimeout(total=30)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(sheet_url) as resp:
                if resp.status != 200:
                    raise RuntimeError(f"Failed to fetch sheet export ({resp.status})")
                payload = await resp.read()
                await asyncio.to_thread(self.sheet_cache_path.write_bytes, payload)
        return self.sheet_cache_path

    async def _load_master_guns(self) -> list[dict[str, Any]]:
        """Load wholesaler source rows using configured sheet name with gid fallback."""
        sheet_path = await self._resolve_sheet_path()
        _, sheet_gid = await self._resolve_sheet_source()
        guns = await asyncio.to_thread(
            self.parse_master_sheet,
            sheet_path,
            config.WHOLESALER_MASTER_SHEET_NAME,
            sheet_gid,
        )
        if guns:
            await gun_catalog_upsert_many(guns)
        return guns

    async def _load_state(self) -> dict[str, Any]:
        wholesale_file = getattr(self, "wholesale_inventory_file", self.state_file)
        store_file = getattr(self, "store_state_file", self.state_file)
        store_inventory_dir = getattr(self, "store_inventory_dir", Path(self.state_file).parent / "inventory" / "stores")

        lots = await wh_lots_get_all()
        stores_db = await wh_stores_get_all()
        shops_db = await wh_shops_get_all()
        settings_db = await wh_settings_get()

        if lots or stores_db or shops_db:
            json_state = await helpers.load_json_file(self.state_file, default={})
            state = {
                "wholesale_lots": lots,
                "stores": stores_db,
                "shop_registry": shops_db,
                "pending_payouts": json_state.get("pending_payouts", []),
                "settings": settings_db,
                "transactions": 0,
            }
            restock = state["settings"].setdefault("restock", {})
            for key, value in self.DEFAULT_RESTOCK_SETTINGS.items():
                restock.setdefault(key, value)
            logger.debug(
                "_load_state: loaded from DB wholesale_lots=%d stores=%d",
                len(state["wholesale_lots"]),
                len(state["stores"]),
            )
            return state

        # Fall back to file loading (first boot / migration not yet run)
        state = await helpers.load_json_file(
            self.state_file,
            default={
                "wholesale_lots": [],
                "transactions": 0,
                "pending_payouts": [],
                "settings": {},
            },
        )

        wholesale_state = await helpers.load_json_file(wholesale_file, default={})

        stores: dict[str, Any] = {}
        shop_registry: dict[str, Any] = {}

        if store_file == self.state_file:
            store_state = state
        else:
            store_state = await helpers.load_json_file(
                store_file,
                default={"stores": {}, "shop_registry": {}},
            )

        legacy_stores = store_state.get("stores", {})
        if isinstance(legacy_stores, dict):
            for store_id, payload in legacy_stores.items():
                if not isinstance(payload, dict):
                    continue
                lots_list = payload.get("lots", [])
                stores[str(store_id)] = {
                    "owner_id": payload.get("owner_id"),
                    "lots": lots_list if isinstance(lots_list, list) else [],
                    "controlled_buyers": payload.get("controlled_buyers", []),
                }

        legacy_registry = store_state.get("shop_registry", {})
        if isinstance(legacy_registry, dict):
            shop_registry.update(legacy_registry)

        if store_inventory_dir.exists():
            for store_file_path in sorted(store_inventory_dir.glob("*.json")):
                payload = await helpers.load_json_file(store_file_path, default={})
                if not isinstance(payload, dict):
                    continue
                store_id = str(payload.get("store_id") or store_file_path.stem)
                lots_list = payload.get("lots", [])
                if isinstance(lots_list, list):
                    owner_id = payload.get("owner_id")
                    controlled = payload.get("controlled_buyers", [])
                    stores[store_id] = {"owner_id": owner_id, "lots": lots_list, "controlled_buyers": controlled}

        state["stores"] = stores
        state["shop_registry"] = shop_registry
        wholesale_lots = wholesale_state.get("wholesale_lots", state.get("wholesale_lots", []))
        state["wholesale_lots"] = wholesale_lots if isinstance(wholesale_lots, list) else []
        state.setdefault("shop_registry", {})
        state.setdefault("stores", {})
        state.setdefault("wholesale_lots", [])
        state.setdefault("pending_payouts", [])
        state.setdefault("settings", {})
        restock = state["settings"].setdefault("restock", {})
        for key, value in self.DEFAULT_RESTOCK_SETTINGS.items():
            restock.setdefault(key, value)
        logger.debug(
            "_load_state: wholesale_lots=%d stores=%d (state_file=%s wholesale_file=%s)",
            len(state["wholesale_lots"]),
            len(state["stores"]),
            self.state_file,
            wholesale_file,
        )
        return state

    @staticmethod
    def _sanitize_positive_int(value: Any, fallback: int) -> int:
        try:
            v = int(value)
            return v if v > 0 else fallback
        except Exception:
            return fallback

    @staticmethod
    def _sanitize_non_negative_int(value: Any, fallback: int) -> int:
        try:
            v = int(value)
            return v if v >= 0 else fallback
        except Exception:
            return fallback

    def _resolve_restock_settings(self, state: dict[str, Any]) -> dict[str, int]:
        raw = state.get("settings", {}).get("restock", {})
        data = {}
        for key, default in self.DEFAULT_RESTOCK_SETTINGS.items():
            if key in {"lots_L", "lots_M", "lots_H"}:
                data[key] = self._sanitize_non_negative_int(raw.get(key), default)
            else:
                data[key] = self._sanitize_positive_int(raw.get(key), default)

        # Keep ranges valid and ensure at least one lot is generated.
        for lvl in ("L", "M", "H"):
            mn_key = f"qty_min_{lvl}"
            mx_key = f"qty_max_{lvl}"
            if data[mn_key] > data[mx_key]:
                data[mn_key], data[mx_key] = data[mx_key], data[mn_key]

        if data["lots_L"] + data["lots_M"] + data["lots_H"] <= 0:
            data["lots_L"] = 1

        data["total_lots"] = max(1, data["total_lots"])

        explicit_type_lots = state.get("settings", {}).get("restock_type_lots", {})
        for weapon_type in self.WEAPON_TYPES:
            for lvl in ("L", "M", "H"):
                key = f"{weapon_type}_{lvl}"
                data[key] = self._sanitize_non_negative_int(explicit_type_lots.get(key), 0)

        return data

    def _generate_restock_lots(
        self,
        guns: list[dict[str, Any]],
        cfg: dict[str, int],
        rng: random.Random,
    ) -> tuple[list[dict[str, Any]], dict[str, int]]:
        guns = [g for g in guns if str(g.get("status", "live")).strip().lower() == "live"]
        by_level = {
            "L": [g for g in guns if g["gun_level"] == "L"],
            "M": [g for g in guns if g["gun_level"] == "M"],
            "H": [g for g in guns if g["gun_level"] == "H"],
        }
        weighted = [
            g
            for g in guns
            for _ in range(self.LEVEL_SETTINGS.get(g["gun_level"], {"weight": 1})["weight"])
        ]

        lots: list[dict[str, Any]] = []
        level_totals = {"L": 0, "M": 0, "H": 0}

        explicit_keys = [f"{weapon_type}_{lvl}" for weapon_type in self.WEAPON_TYPES for lvl in ("L", "M", "H")]
        has_explicit_mix = any(cfg.get(k, 0) > 0 for k in explicit_keys)

        if has_explicit_mix:
            for weapon_type in self.WEAPON_TYPES:
                for requested_level in ("L", "M", "H"):
                    target = int(cfg.get(f"{weapon_type}_{requested_level}", 0))
                    if target <= 0:
                        continue
                    pool = [
                        g for g in guns
                        if g.get("weapon_type") == weapon_type and g.get("gun_level") == requested_level
                    ]
                    if not pool:
                        pool = [g for g in by_level[requested_level] if g.get("weapon_type") == weapon_type]
                    if not pool:
                        continue

                    for _ in range(target):
                        gun = rng.choice(pool)
                        qty = rng.randint(cfg[f"qty_min_{requested_level}"], cfg[f"qty_max_{requested_level}"])
                        lots.append(
                            {
                                "lot_id": f"lot-{datetime.now(timezone.utc).strftime('%Y%m%d')}-{uuid.uuid4().hex[:6]}",
                                "gun_name": gun["gun_name"],
                                "gun_level": requested_level,
                                "gun_category": gun.get("gun_category") or "",
                                "weapon_type": gun.get("weapon_type") or "",
                                "unit_cost": int(gun["price_new"]),
                                "qty_available": qty,
                                "restriction": gun.get("restriction", "basic"),
                                "created_at": self._now_iso(),
                            }
                        )
                        level_totals[requested_level] += qty
        else:

            for requested_level in ("L", "M", "H"):
                target = cfg[f"lots_{requested_level}"]
                pool = by_level[requested_level] if by_level[requested_level] else weighted
                if not pool or target <= 0:
                    continue

                for _ in range(target):
                    gun = rng.choice(pool)
                    actual_level = str(gun.get("gun_level", requested_level))
                    if actual_level not in {"L", "M", "H"}:
                        actual_level = requested_level
                    qty = rng.randint(cfg[f"qty_min_{actual_level}"], cfg[f"qty_max_{actual_level}"])
                    lots.append(
                        {
                            "lot_id": f"lot-{datetime.now(timezone.utc).strftime('%Y%m%d')}-{uuid.uuid4().hex[:6]}",
                            "gun_name": gun["gun_name"],
                            "gun_level": actual_level,
                            "gun_category": gun.get("gun_category") or "",
                            "weapon_type": gun.get("weapon_type") or "",
                            "unit_cost": int(gun["price_new"]),
                            "qty_available": qty,
                            "restriction": gun.get("restriction", "basic"),
                            "created_at": self._now_iso(),
                        }
                    )
                    level_totals[actual_level] += qty

        if len(lots) > cfg["total_lots"]:
            lots = rng.sample(lots, cfg["total_lots"])

        merged: dict[str, dict[str, Any]] = {}
        for lot in lots:
            key = f"{lot['gun_name']}|{lot['gun_level']}|{lot['unit_cost']}|{lot.get('weapon_type', '')}|{lot.get('gun_category', '')}"
            if key in merged:
                merged[key]["qty_available"] += lot["qty_available"]
            else:
                merged[key] = dict(lot)
        lots = list(merged.values())

        level_totals = {
            "L": sum(int(lot.get("qty_available", 0)) for lot in lots if lot.get("gun_level") == "L"),
            "M": sum(int(lot.get("qty_available", 0)) for lot in lots if lot.get("gun_level") == "M"),
            "H": sum(int(lot.get("qty_available", 0)) for lot in lots if lot.get("gun_level") == "H"),
        }

        return lots, level_totals

    async def _save_state(self, state: dict[str, Any]) -> bool:
        wholesale_file = Path(getattr(self, "wholesale_inventory_file", self.state_file))
        store_file = Path(getattr(self, "store_state_file", self.state_file))
        default_data_dir = Path(getattr(self, "data_dir", Path(self.state_file).parent))
        store_inventory_dir = Path(getattr(self, "store_inventory_dir", default_data_dir / "inventory" / "stores"))

        # Keep a compatibility snapshot in the main state file so deployments
        # that only persist `state.json` (and not the split inventory files)
        # still retain stock across restarts.
        state_main = dict(state)

        wholesale_payload = {
            "wholesale_lots": state.get("wholesale_lots", []),
            "updated_at": self._now_iso(),
        }

        stores_payload = {
            "shop_registry": state.get("shop_registry", {}),
            "stores": {},
            "updated_at": self._now_iso(),
        }

        main_ok = await helpers.save_json_file(self.state_file, state_main)
        wholesale_ok = await helpers.save_json_file(wholesale_file, wholesale_payload)

        store_index_ok = True

        store_inventory_dir.mkdir(parents=True, exist_ok=True)
        stores = state.get("stores", {})
        desired_paths: set[Path] = set()
        store_files_ok = True
        if isinstance(stores, dict):
            for store_id, store_data in stores.items():
                dest = self._store_inventory_file(str(store_id))
                desired_paths.add(dest)
                stores_payload["stores"][store_id] = {
                    "owner_id": store_data.get("owner_id"),
                    "inventory_file": str(dest),
                }
                payload = {
                    "store_id": store_id,
                    "owner_id": store_data.get("owner_id"),
                    "lots": store_data.get("lots", []),
                    "controlled_buyers": store_data.get("controlled_buyers", []),
                    "updated_at": self._now_iso(),
                }
                if not await helpers.save_json_file(dest, payload):
                    store_files_ok = False

        for old_file in self._list_store_inventory_files():
            if old_file not in desired_paths:
                try:
                    old_file.unlink(missing_ok=True)
                except Exception:
                    logger.exception("Failed to remove stale store inventory file: %s", old_file)
                    store_files_ok = False

        if store_file != self.state_file:
            store_index_ok = await helpers.save_json_file(store_file, stores_payload)

        db_lots_ok = await wh_lots_replace_all(state.get("wholesale_lots", []))
        db_stores_ok = await wh_stores_replace_all(state.get("stores", {}))
        db_shops_ok = await wh_shops_replace_all(state.get("shop_registry", {}))
        db_settings_ok = await wh_settings_save(state.get("settings", {}))
        db_ok = db_lots_ok and db_stores_ok and db_shops_ok and db_settings_ok
        ok = main_ok and wholesale_ok and store_index_ok and store_files_ok and db_ok
        if not ok:
            logger.error(
                "Wholesaler persistence failure main_ok=%s wholesale_ok=%s store_index_ok=%s store_files_ok=%s "
                "state_file=%s wholesale_file=%s store_file=%s store_inventory_dir=%s",
                main_ok,
                wholesale_ok,
                store_index_ok,
                store_files_ok,
                self.state_file,
                wholesale_file,
                store_file,
                store_inventory_dir,
            )
        return ok

    async def _append_tx(self, tx: dict[str, Any]) -> bool:
        file_ok = await helpers.append_json_file(self.tx_file, tx)
        db_ok = await wh_tx_append(tx)
        return file_ok and db_ok

    @staticmethod
    def _coerce_role_ids(value: Any) -> set[int]:
        """Normalize role-id config values from int/list/CSV formats."""
        if value is None:
            return set()
        if isinstance(value, (int, float)):
            return {int(value)}
        if isinstance(value, str):
            parts = [part.strip() for part in value.split(",")]
            return {
                int(part)
                for part in parts
                if part and part.lstrip("-").isdigit()
            }

        role_ids: set[int] = set()
        try:
            iterator = iter(value)
        except TypeError:
            return set()

        for item in iterator:
            if isinstance(item, (int, float)):
                role_ids.add(int(item))
                continue
            text = str(item).strip()
            if text and text.lstrip("-").isdigit():
                role_ids.add(int(text))
        return role_ids

    def _is_admin(self, member: discord.Member) -> bool:
        admin_role_ids = self._coerce_role_ids(getattr(config, "WHOLESALER_ADMIN_ROLE_IDS", []))
        return bool(member.guild_permissions.administrator) or any(r.id in admin_role_ids for r in member.roles)

    def _is_store_owner(self, member: discord.Member) -> bool:
        store_role_ids = self._coerce_role_ids(getattr(config, "WHOLESALER_STORE_ROLE_IDS", []))
        return any(r.id in store_role_ids for r in member.roles)

    async def _get_audit_channel(self) -> Optional[discord.TextChannel]:
        """Resolve the GUN_LOG_CHANNEL_ID to a channel object, or return None."""
        channel_id = int(getattr(config, "GUN_LOG_CHANNEL_ID", 0) or 0)
        if channel_id <= 0:
            logger.warning("Missing gun-log audit channel id=%s", channel_id)
            return None
        channel = self.bot.get_channel(channel_id)
        if channel is None:
            try:
                channel = await self.bot.fetch_channel(channel_id)
            except Exception:
                logger.warning("Missing gun-log audit channel id=%s", channel_id)
                return None
        if not channel:
            logger.warning("Missing gun-log audit channel id=%s", channel_id)
            return None
        return channel

    async def _audit_send(self, text: str) -> None:
        """Send a plain operational audit event as a minimal embed."""
        channel = await self._get_audit_channel()
        if channel is None:
            return
        # Parse the action tag from the leading [ACTION] token if present
        action = "Gun Shop Audit"
        body = text
        if text.startswith("["):
            end = text.find("]")
            if end != -1:
                action = text[1:end].replace("_", " ").title()
                body = text[end + 1:].strip()
        embed = discord.Embed(
            title=f"🔔 {action}",
            description=f"```{body}```" if body else None,
            color=discord.Color.dark_grey(),
            timestamp=datetime.now(timezone.utc),
        )
        embed.set_footer(text="NightCityBot Audit Log")
        try:
            await channel.send(embed=embed, allowed_mentions=discord.AllowedMentions.none())
        except Exception:
            logger.exception("Failed to send gun-log audit line")

    async def _audit_embed_send(self, embed: discord.Embed) -> None:
        """Send a pre-built structured audit embed to the gun-log channel."""
        channel = await self._get_audit_channel()
        if channel is None:
            return
        try:
            await channel.send(embed=embed, allowed_mentions=discord.AllowedMentions.none())
        except Exception:
            logger.exception("Failed to send gun-log audit embed")


    async def _system_enabled(self, ctx: commands.Context) -> bool:
        control = self.bot.get_cog("SystemControl")
        if control and not control.is_enabled("gun_shop"):
            await ctx.send("⚠️ The gun shop system is currently disabled.")
            return False
        return True

    async def _ensure_member(self, ctx: commands.Context) -> Optional[discord.Member]:
        if not ctx.guild or not isinstance(ctx.author, discord.Member):
            await ctx.send("❌ This command can only be used in the server.")
            return None
        return ctx.author

    async def _resolve_member(self, ctx: commands.Context, raw) -> Optional[discord.Member]:
        if isinstance(raw, discord.Member):
            return raw
        if not ctx.guild:
            return None
        raw = str(raw)
        match = re.match(r"<@!?(\d+)>", raw)
        user_id = int(match.group(1)) if match else None
        if not user_id and raw.isdigit():
            user_id = int(raw)
        if user_id:
            member = ctx.guild.get_member(user_id)
            if member:
                return member
            try:
                return await ctx.guild.fetch_member(user_id)
            except (discord.NotFound, discord.HTTPException):
                return None
        try:
            return await commands.MemberConverter().convert(ctx, raw)
        except commands.MemberNotFound:
            return None

    @staticmethod
    def _store_id(guild_id: int, owner_id: int) -> str:
        return f"{guild_id}:{owner_id}"

    def _shop_display_name(
        self,
        state: dict[str, Any],
        owner_id: int,
        requested_shop: Optional[str] = None,
    ) -> str:
        if requested_shop:
            return self._normalize_shop_name(requested_shop)

        registry = state.get("shop_registry", {})
        aliases = sorted(name for name, mapped_owner in registry.items() if int(mapped_owner) == int(owner_id))
        if aliases:
            return aliases[0]

        return f"owner:{owner_id}"

    def _build_tx(self, tx_type: str, **kwargs: Any) -> dict[str, Any]:
        return {
            "tx_id": f"tx-{uuid.uuid4().hex[:12]}",
            "type": tx_type,
            "timestamp": self._now_iso(),
            "status": "SUCCESS",
            "error_details": "",
            **kwargs,
        }

    async def _resolve_store_owner_id(
        self, ctx: commands.Context, state: dict[str, Any], shop_or_mention: Optional[str], default_owner: int
    ) -> int:
        if not shop_or_mention:
            return default_owner
        if shop_or_mention.startswith("<@") and shop_or_mention.endswith(">"):
            return int(shop_or_mention.strip("<@!>"))
        key = self._normalize_shop_name(shop_or_mention)
        return int(state.get("shop_registry", {}).get(key, default_owner))

    async def _get_total_balance(self, user_id: int) -> Optional[tuple[int, int, int]]:
        balance = await self.unbelievaboat.get_balance(user_id)
        if not balance:
            return None
        cash = int(balance.get("cash", 0))
        bank = int(balance.get("bank", 0))
        return cash, bank, cash + bank

    async def _deduct_funds(self, user_id: int, amount: int, reason: str) -> tuple[bool, str]:
        balances = await self._get_total_balance(user_id)
        if balances is None:
            return False, "Unable to fetch user balance"
        cash, _bank, total = balances
        if total < amount:
            return False, f"Insufficient funds (${total}/${amount})"

        cash_deduct = min(max(cash, 0), amount)
        bank_deduct = max(0, amount - cash_deduct)
        payload: dict[str, int] = {}
        if cash_deduct:
            payload["cash"] = -cash_deduct
        if bank_deduct:
            payload["bank"] = -bank_deduct
        ok = await self.unbelievaboat.update_balance(user_id, payload, reason=reason)
        return (ok, "" if ok else "Failed to deduct funds")

    async def _credit_funds(self, user_id: int, amount: int, reason: str) -> bool:
        logger.info("_credit_funds: crediting user=%s amount=%d reason=%s", user_id, amount, reason)
        result = await self.unbelievaboat.update_balance(user_id, {"cash": amount}, reason=reason)
        if not result:
            logger.error("_credit_funds: FAILED to credit user=%s amount=%d", user_id, amount)
        return result

    async def _request_admin_approval(
        self,
        ctx: commands.Context,
        seller: discord.Member,
        buyer: discord.Member,
        lot: dict,
        qty: int,
        total_price: int,
        character_name: str,
    ) -> bool:
        """Post a restricted-sale approval request and wait for an admin reaction."""
        channel_id = int(getattr(config, "GUN_LOG_CHANNEL_ID", 0) or 0)
        if channel_id <= 0:
            await ctx.send("❌ Gun-log channel not configured — cannot process restricted sales.")
            return False

        channel = self.bot.get_channel(channel_id)
        if channel is None:
            try:
                channel = await self.bot.fetch_channel(channel_id)
            except Exception:
                await ctx.send("❌ Gun-log channel not found — cannot process restricted sales.")
                return False

        embed = discord.Embed(
            title="🔒 Restricted Sale — Admin Approval Required",
            color=discord.Color.orange(),
        )
        embed.add_field(name="Seller", value=seller.mention, inline=True)
        embed.add_field(name="Buyer", value=f"{buyer.mention} (character: {character_name})", inline=True)
        embed.add_field(name="Item", value=f"{qty}x {lot['gun_name']} (Tier {lot['gun_level']})", inline=True)
        embed.add_field(name="Total Price", value=f"${total_price}", inline=True)
        embed.add_field(name="Lot ID", value=f"`{lot['lot_id']}`", inline=True)
        embed.set_footer(text="React ✅ to approve or ❌ to deny. Expires in 5 minutes.")

        try:
            msg = await channel.send(embed=embed)
            await msg.add_reaction("✅")
            await msg.add_reaction("❌")
        except Exception:
            logger.exception("Failed to post restricted sale approval request")
            await ctx.send("❌ Failed to post approval request to audit channel.")
            return False

        await ctx.send(f"⏳ Restricted sale pending admin approval in the audit channel. Waiting up to 5 minutes...")

        def check(reaction, user):
            return (
                reaction.message.id == msg.id
                and str(reaction.emoji) in ("✅", "❌")
                and user != self.bot.user
                and self._is_admin(user)
            )

        try:
            reaction, admin_user = await self.bot.wait_for("reaction_add", timeout=300.0, check=check)
        except asyncio.TimeoutError:
            await ctx.send("❌ Restricted sale timed out — no admin response within 5 minutes.")
            try:
                await msg.edit(embed=embed.set_footer(text="EXPIRED — no admin response."))
            except Exception:
                logger.warning("Suppressed exception", exc_info=True)
            return False

        if str(reaction.emoji) == "✅":
            await ctx.send(f"✅ Restricted sale approved by {admin_user.mention}. Processing...")
            try:
                await msg.edit(embed=embed.set_footer(text=f"APPROVED by {admin_user.display_name}"))
            except Exception:
                logger.warning("Suppressed exception", exc_info=True)
            _approved_embed = discord.Embed(
                title="✅ Restricted Sale Approved",
                color=discord.Color.green(),
                timestamp=datetime.now(timezone.utc),
            )
            _approved_embed.add_field(name="Admin", value=f"{admin_user.mention} ({admin_user.display_name})", inline=True)
            _approved_embed.add_field(name="Seller", value=f"{seller.mention} ({seller.display_name})", inline=True)
            _approved_embed.add_field(name="Buyer", value=f"{buyer.mention} ({buyer.display_name})", inline=True)
            _approved_embed.add_field(name="Item", value=f"{qty}x {lot['gun_name']}", inline=True)
            _approved_embed.add_field(name="Price", value=f"${total_price:,}", inline=True)
            _approved_embed.set_footer(text="NightCityBot Audit Log")
            await self._audit_embed_send(_approved_embed)
            return True
        else:
            await ctx.send(f"❌ Restricted sale denied by {admin_user.mention}.")
            try:
                await msg.edit(embed=embed.set_footer(text=f"DENIED by {admin_user.display_name}"))
            except Exception:
                logger.warning("Suppressed exception", exc_info=True)
            _denied_embed = discord.Embed(
                title="❌ Restricted Sale Denied",
                color=discord.Color.red(),
                timestamp=datetime.now(timezone.utc),
            )
            _denied_embed.add_field(name="Admin", value=f"{admin_user.mention} ({admin_user.display_name})", inline=True)
            _denied_embed.add_field(name="Seller", value=f"{seller.mention} ({seller.display_name})", inline=True)
            _denied_embed.add_field(name="Buyer", value=f"{buyer.mention} ({buyer.display_name})", inline=True)
            _denied_embed.add_field(name="Item", value=f"{qty}x {lot['gun_name']}", inline=True)
            _denied_embed.add_field(name="Price", value=f"${total_price:,}", inline=True)
            _denied_embed.set_footer(text="NightCityBot Audit Log")
            await self._audit_embed_send(_denied_embed)
            return False

    async def auto_refresh_weekly_after_cyberware(self) -> bool:
        """Auto-restock once per week, called by cyberware weekly process."""
        return await self._auto_restock_if_due(datetime.now(timezone.utc), trigger="CYBERWARE")

    async def _auto_restock_if_due(self, now: datetime, trigger: str) -> bool:
        """Run one weekly restock if we have not yet refreshed this Sunday."""
        control = self.bot.get_cog("SystemControl")
        if control and not control.is_enabled("gun_shop"):
            return False
        try:
            guns = await self._load_master_guns()
        except Exception:
            logger.exception("Auto wholesaler refresh failed during sheet read")
            return False

        guns = [g for g in guns if str(g.get("status", "live")).strip().lower() == "live"]
        if not guns:
            return False

        sunday_key = now.strftime("%Y-%m-%d")
        async with self.lock:
            state = await self._load_state()
            settings = state.setdefault("settings", {})
            last_key = str(settings.get("last_auto_restock_sunday", ""))
            if last_key == sunday_key:
                return True

            # Legacy key fallback to avoid duplicate restock in the same week.
            legacy_week_key = now.strftime("%Y-W%U")
            if str(settings.get("last_auto_restock_week", "")) == legacy_week_key:
                settings["last_auto_restock_sunday"] = sunday_key
                await self._save_state(state)
                return True

            cfg = self._resolve_restock_settings(state)
            rng = random.Random()
            lots, _level_totals = self._generate_restock_lots(guns, cfg, rng)

            state["wholesale_lots"] = lots
            settings["last_auto_restock_sunday"] = sunday_key
            settings["last_auto_restock_week"] = legacy_week_key
            saved = await self._save_state(state)
            if not saved:
                logger.error("Auto restock save failed trigger=%s", trigger)
            await gun_catalog_sync_qty_from_lots(lots)

        await self._audit_send(
            f"[WHOLESALE_AUTO_RESTOCK] trigger={trigger} lots={len(lots)} sunday={sunday_key}"
        )
        return True

