"""Cyberware shop sheet parsing and download utilities."""
import asyncio
import logging
import re
from pathlib import Path
from typing import Any, Optional
from urllib.parse import parse_qs, urlparse

import aiohttp
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _normalize_sheet_url(raw: str) -> str:
    """Convert a Google Sheets share URL to an XLSX export URL."""
    raw = (raw or "").strip()
    if raw.startswith("<") and raw.endswith(">"):
        raw = raw[1:-1].strip()
    if not raw:
        return raw
    parsed = urlparse(raw)
    if "docs.google.com" not in parsed.netloc:
        return raw
    if "/export" in parsed.path:
        return raw
    match = re.search(r"/spreadsheets/(?:u/\d+/)?d/([a-zA-Z0-9-_]+)", parsed.path)
    if not match:
        return raw
    sheet_id = match.group(1)
    gid = parse_qs(parsed.query).get("gid", [None])[0]
    if not gid and parsed.fragment:
        frag_qs = parse_qs(parsed.fragment)
        gid = frag_qs.get("gid", [None])[0]
        if not gid and parsed.fragment.startswith("gid="):
            gid = parsed.fragment.split("=", 1)[1].split("&", 1)[0]
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    if gid:
        export_url += f"&gid={gid}"
    return export_url


async def download_sheet(url: str, save_path: Path) -> None:
    """Download a Google Sheet as XLSX and save to save_path."""
    export_url = _normalize_sheet_url(url)
    timeout = aiohttp.ClientTimeout(total=30)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        async with session.get(export_url) as resp:
            if resp.status != 200:
                raise RuntimeError(
                    f"Failed to download cyberware sheet ({resp.status}): {export_url}"
                )
            payload = await resp.read()
    save_path.parent.mkdir(parents=True, exist_ok=True)
    await asyncio.to_thread(save_path.write_bytes, payload)


def parse_cyberware_sheet(xlsx_path: Path | str) -> list[dict[str, Any]]:
    """Parse the first tab of the XLSX for item name, price, CWP, and description.

    Returns a list of dicts with keys ``name``, ``price`` (int), ``cwp`` (str),
    and ``description`` (str).  Skips rows without a name or a positive price.
    """
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    row_iter = ws.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    if not header_row:
        wb.close()
        return []

    header = [str(c).strip() if c is not None else "" for c in header_row]
    header_lower = {h.lower(): i for i, h in enumerate(header)}

    def _find_col(candidates: list[str], fallback: int) -> int:
        for name in candidates:
            idx = header_lower.get(name.lower())
            if idx is not None:
                return idx
        return fallback

    name_idx = _find_col(
        ["item name", "name", "cyberware", "cyberware name", "item", "part name"],
        0,
    )
    price_idx = _find_col(
        [
            "pricing", "price", "cost", "selling price", "sell price",
            "price (eb)", "cost (eb)", "price new", "price (new)",
        ],
        1,
    )
    wholesale_price_idx = _find_col(
        [
            "wholesale price", "wholesale", "price (wholesale)",
            "wholesale_price", "wholesale cost", "buy in",
        ],
        -1,
    )
    cwp_idx = _find_col(
        ["cwp", "cyberware points", "cw points", "cyber points", "cwp cost", "cwp value"],
        -1,
    )
    desc_idx = _find_col(
        ["description", "desc", "effect", "effects", "details", "notes", "ability", "abilities"],
        -1,
    )
    slot_idx = _find_col(
        ["slot", "body slot", "body location", "location"],
        -1,
    )

    logger.info(
        "parse_cyberware_sheet: headers=%s | name_col=%d (%s) | price_col=%d (%s) | cwp_col=%d | desc_col=%d | slot_col=%d",
        header,
        name_idx, header[name_idx] if name_idx < len(header) else "?",
        price_idx, header[price_idx] if price_idx < len(header) else "?",
        cwp_idx,
        desc_idx,
        slot_idx,
    )

    items: list[dict[str, Any]] = []
    for row in row_iter:
        if not row:
            continue
        raw_name = row[name_idx] if name_idx < len(row) else None
        raw_price = row[price_idx] if price_idx < len(row) else None

        if raw_name is None:
            continue
        item_name = str(raw_name).strip()
        if not item_name:
            continue

        price = _to_int(raw_price)
        if price is None or price <= 0:
            continue

        wholesale_price = price
        if wholesale_price_idx >= 0 and wholesale_price_idx < len(row) and row[wholesale_price_idx] is not None:
            wp = _to_int(row[wholesale_price_idx])
            if wp is not None and wp > 0:
                wholesale_price = wp

        cwp = ""
        if cwp_idx >= 0 and cwp_idx < len(row) and row[cwp_idx] is not None:
            cwp = str(row[cwp_idx]).strip()

        description = ""
        if desc_idx >= 0 and desc_idx < len(row) and row[desc_idx] is not None:
            description = str(row[desc_idx]).strip()

        slot = ""
        if slot_idx >= 0 and slot_idx < len(row) and row[slot_idx] is not None:
            slot = str(row[slot_idx]).strip()

        items.append({"name": item_name, "price": price, "wholesale_price": wholesale_price, "cwp": cwp, "description": description, "slot": slot})

    wb.close()
    return items


def _to_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    stripped = str(value).replace(",", "").replace("$", "").strip()
    if not stripped:
        return None
    try:
        return int(float(stripped))
    except ValueError:
        return None
